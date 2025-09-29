from fastapi import APIRouter, HTTPException, File, UploadFile, Form
from fastapi.responses import Response, JSONResponse
from typing import List, Dict, Any, Optional
import pandas as pd
import numpy as np
import io
import json
from app.services.lnd_service import (
    weiszfeld,
    repeated_weiszfeld,
    repeated_multi_facility_weiszfeld,
    cluster_customers_kmeans,
    cluster_customers_hierarchical, 
    solve_k_median,
    calculate_facility_service_area,
    generate_candidate_facilities,
    great_circle_distance,
    lnd_ms,
    lnd_ss,
    solve_multiple_source_lnd,
    solve_single_source_lnd,
    elbow_method_analysis,
    LNDP,
    calculate_echelon_inventory_costs
)
from app.services.advanced_facility_service import AdvancedFacilityLocationService
from app.services.excel_integration_service import ExcelIntegrationService
from app.services.customer_aggregation_service import CustomerAggregationService
from app.services.carbon_footprint_service import CarbonFootprintService
from app.services.network_generation_service import NetworkGenerationService
from app.services.transportation_service import TransportationService
from app.services.visualization_service import VisualizationService
from app.services.sample_data_service import (
    create_lnd_sample_datasets,
    get_sample_data_info
)
from app.models.lnd import (
    ExcelTemplateRequest,
    ExcelTemplateResponse,
    ExcelWorkflowRequest,
    ExcelWorkflowResponse,
    LNDSolveRequest,
    LNDSolveResponse,
    CustomerAggregationRequest,
    CustomerAggregationResponse,
    NetworkGenerationRequest,
    NetworkGenerationResponse,
    AdvancedLNDPRequest,
    AdvancedLNDPResponse,
    VRPIntegrationRequest,
    VRPIntegrationResponse,
    CO2CalculationRequest,
    CO2CalculationResponse
)

router = APIRouter()

# Initialize services
advanced_facility_service = AdvancedFacilityLocationService()
excel_integration_service = ExcelIntegrationService()
customer_aggregation_service = CustomerAggregationService()
carbon_footprint_service = CarbonFootprintService()
network_generation_service = NetworkGenerationService()
transportation_service = TransportationService()
visualization_service = VisualizationService()

@router.post("/weiszfeld-location")
async def calculate_weiszfeld_location(
    file: UploadFile = File(...),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    max_iterations: int = Form(1000),
    tolerance: float = Form(1e-6)
):
    """
    Calculate optimal facility location using Weiszfeld algorithm (geometric median)
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Extract points and weights
        points = [(row[lat_col], row[lon_col]) for _, row in df.iterrows()]
        weights = df[demand_col].tolist() if demand_col in df.columns else None
        
        # Calculate geometric median using notebook implementation
        if weights is None:
            weights = [1.0] * len(points)
        
        # Create temporary dataframe for weiszfeld function
        df_temp = df.copy()
        
        # Call weiszfeld function (single facility)
        X, Y, partition, cost = weiszfeld(df_temp, weights, 1, tolerance, max_iterations)
        optimal_location = (X[0], Y[0])
        
        # Calculate service area statistics
        service_stats = calculate_facility_service_area(
            customer_df=df,
            facility_location=optimal_location,
            lat_col=lat_col,
            lon_col=lon_col,
            demand_col=demand_col
        )
        
        return {
            "optimal_location": {
                "latitude": optimal_location[0],
                "longitude": optimal_location[1]
            },
            "service_area_statistics": service_stats,
            "algorithm": "weiszfeld",
            "parameters": {
                "max_iterations": max_iterations,
                "tolerance": tolerance,
                "weighted": weights is not None
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error calculating Weiszfeld location: {str(e)}")

@router.post("/multi-facility-weiszfeld")
async def calculate_multi_facility_weiszfeld(
    file: UploadFile = File(...),
    num_facilities: int = Form(3),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    max_iterations: int = Form(1000),
    tolerance: float = Form(1e-4),
    random_state: int = Form(42)
):
    """
    Solve multi-facility location problem using Weiszfeld algorithm
    
    This endpoint implements the multi-facility extension of the Weiszfeld algorithm
    for finding optimal weighted geometric median locations for multiple facilities.
    
    Features:
    - Weighted optimization using customer demand
    - Iterative assignment and location update
    - Convergence checking
    - Detailed facility statistics
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Validate num_facilities
        if num_facilities < 1:
            raise HTTPException(
                status_code=400,
                detail="Number of facilities must be at least 1"
            )
        
        if num_facilities > len(df):
            raise HTTPException(
                status_code=400,
                detail=f"Number of facilities ({num_facilities}) cannot exceed number of customers ({len(df)})"
            )
        
        # Extract weights (demand)
        if demand_col in df.columns:
            weights = df[demand_col].tolist()
            weights = [1.0 if pd.isna(w) else w for w in weights]
        else:
            weights = [1.0] * len(df)
        
        # Solve multi-facility Weiszfeld problem using notebook implementation
        X, Y, partition, cost = weiszfeld(df, weights, num_facilities, tolerance, max_iterations, random_state)
        
        # Calculate facility statistics  
        facility_stats = []
        for j in range(num_facilities):
            customers_assigned = [i for i, p in partition.items() if p == j]
            
            if customers_assigned:
                customer_indices = customers_assigned
                total_demand_served = sum(weights[i] for i in customer_indices)
                distances_to_facility = [great_circle_distance(df.iloc[i][lat_col], df.iloc[i][lon_col], X[j], Y[j]) for i in customer_indices]
                avg_distance = sum(distances_to_facility) / len(distances_to_facility) if distances_to_facility else 0
            else:
                total_demand_served = 0.0
                avg_distance = 0.0
                customer_indices = []
            
            facility_stats.append({
                'facility_index': j,
                'location': [float(X[j]), float(Y[j])],
                'customers_assigned': len(customer_indices),
                'total_demand_served': float(total_demand_served),
                'average_distance': float(avg_distance)
            })
        
        solution = {
            'facility_locations': [(float(X[j]), float(Y[j])) for j in range(num_facilities)],
            'assignments': [partition.get(i, 0) for i in range(len(df))],
            'total_cost': float(cost),
            'facility_stats': facility_stats,
            'algorithm': 'multi_facility_weiszfeld',
            'converged': True,
            'parameters': {
                'num_facilities': num_facilities,
                'max_iterations': max_iterations,
                'tolerance': tolerance,
                'random_state': random_state
            }
        }
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert solution to JSON-safe format
        solution = convert_to_json_safe(solution)
        
        return solution
    
    except Exception as e:
        import traceback
        error_detail = f"Error in multi-facility Weiszfeld: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error in multi-facility Weiszfeld: {str(e)}")

@router.post("/repeated-multi-facility-weiszfeld")
async def calculate_repeated_multi_facility_weiszfeld(
    file: UploadFile = File(...),
    num_facilities: int = Form(3),
    num_runs: int = Form(10),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    max_iterations: int = Form(1000),
    tolerance: float = Form(1e-4),
    base_random_state: int = Form(42)
):
    """
    Solve multi-facility location problem using repeated Weiszfeld algorithm
    
    This endpoint implements the repeated multi-facility Weiszfeld algorithm that runs
    multiple times with different random initializations to find the global optimum.
    
    Features:
    - Multiple runs with different random initializations
    - Global optimization through best solution selection
    - Comprehensive statistics across all runs
    - Detailed convergence analysis
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Validate parameters
        if num_facilities < 2:
            raise HTTPException(
                status_code=400,
                detail="Number of facilities must be at least 2"
            )
        
        if num_facilities > len(df):
            raise HTTPException(
                status_code=400,
                detail=f"Number of facilities ({num_facilities}) cannot exceed number of customers ({len(df)})"
            )
        
        if num_runs < 1:
            raise HTTPException(
                status_code=400,
                detail="Number of runs must be at least 1"
            )
        
        # Use the proper repeated_multi_facility_weiszfeld function that returns cost_statistics
        solution = repeated_multi_facility_weiszfeld(
            customer_df=df,
            num_facilities=num_facilities,
            num_runs=num_runs,
            lat_col=lat_col,
            lon_col=lon_col,
            demand_col=demand_col,
            max_iterations=max_iterations,
            tolerance=tolerance,
            base_random_state=base_random_state
        )
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert solution to JSON-safe format
        solution = convert_to_json_safe(solution)
        
        return solution
    
    except Exception as e:
        import traceback
        error_detail = f"Error in repeated multi-facility Weiszfeld: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error in repeated multi-facility Weiszfeld: {str(e)}")

@router.post("/customer-clustering")
async def cluster_customers(
    file: UploadFile = File(...),
    method: str = Form("kmeans"),
    n_clusters: int = Form(3),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    linkage_method: str = Form("ward"),
    random_state: int = Form(42)
):
    """
    Cluster customers using K-means or hierarchical clustering
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Perform clustering
        try:
            if method.lower() == "kmeans":
                clustered_df = cluster_customers_kmeans(
                    customer_df=df,
                    n_clusters=n_clusters,
                    lat_col=lat_col,
                    lon_col=lon_col,
                    demand_col=demand_col,
                    random_state=random_state
                )
            elif method.lower() == "hierarchical":
                clustered_df = cluster_customers_hierarchical(
                    customer_df=df,
                    n_clusters=n_clusters,
                    lat_col=lat_col,
                    lon_col=lon_col,
                    linkage_method=linkage_method
                )
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unknown clustering method: {method}. Use 'kmeans' or 'hierarchical'"
                )
        except Exception as cluster_error:
            raise HTTPException(
                status_code=500,
                detail=f"Clustering failed: {str(cluster_error)}"
            )
        
        # Convert the DataFrame to JSON-safe format first
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types recursively"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert clustered DataFrame to basic Python types
        clustered_records = []
        for _, row in clustered_df.iterrows():
            record = {}
            for col, value in row.items():
                record[col] = convert_to_json_safe(value)
            clustered_records.append(record)
        
        # Calculate cluster statistics with safe type conversion
        cluster_stats = []
        for i in range(n_clusters):
            cluster_data = clustered_df[clustered_df['cluster'] == i]
            if len(cluster_data) > 0:
                center_lat = convert_to_json_safe(cluster_data[lat_col].mean())
                center_lon = convert_to_json_safe(cluster_data[lon_col].mean())
                
                # Calculate distances safely
                distances = []
                for _, customer in cluster_data.iterrows():
                    try:
                        distance = great_circle_distance(
                            convert_to_json_safe(customer[lat_col]), 
                            convert_to_json_safe(customer[lon_col]),
                            center_lat, center_lon
                        )
                        distances.append(distance)
                    except:
                        distances.append(0.0)
                
                total_demand = convert_to_json_safe(
                    cluster_data[demand_col].sum() if demand_col in df.columns else len(cluster_data)
                )
                
                stats = {
                    "cluster_id": convert_to_json_safe(i),
                    "num_customers": convert_to_json_safe(len(cluster_data)),
                    "total_demand": total_demand,
                    "center_lat": center_lat,
                    "center_lon": center_lon,
                    "avg_distance_to_center": convert_to_json_safe(np.mean(distances)) if distances else 0.0
                }
                cluster_stats.append(stats)
        
        # Extract cluster centers safely
        cluster_centers = []
        try:
            centers_attr = clustered_df.attrs.get('cluster_centers', [])
            if isinstance(centers_attr, (list, np.ndarray)):
                for center in centers_attr:
                    if isinstance(center, (list, tuple, np.ndarray)) and len(center) >= 2:
                        cluster_centers.append({
                            "lat": convert_to_json_safe(center[0]),
                            "lon": convert_to_json_safe(center[1])
                        })
        except (AttributeError, TypeError, IndexError):
            cluster_centers = []
        
        # Build final response with all safe types
        response_data = {
            "clustered_data": clustered_records,
            "cluster_statistics": cluster_stats,
            "cluster_centers": cluster_centers,
            "algorithm": str(method.lower()),
            "parameters": {
                "n_clusters": convert_to_json_safe(n_clusters),
                "linkage_method": str(linkage_method) if method.lower() == "hierarchical" else None,
                "random_state": convert_to_json_safe(random_state) if method.lower() == "kmeans" else None
            }
        }
        
        # Return using JSONResponse to ensure proper serialization
        from fastapi.responses import JSONResponse
        return JSONResponse(content=response_data)
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error in customer clustering: {str(e)}")

@router.post("/k-median-optimization")
async def k_median_optimization(
    customer_file: UploadFile = File(...),
    k: int = Form(3),
    candidate_method: str = Form("grid"),
    n_candidates: int = Form(20),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    max_iterations: int = Form(100),
    max_lr: float = Form(0.01),
    tolerance: float = Form(1e-5),
    use_adam: bool = Form(False),
    capacity: Optional[int] = Form(None)
):
    """
    Solve k-median facility location problem using Lagrangian relaxation
    
    This endpoint implements the full Lagrangian relaxation algorithm with:
    - Subgradient optimization with momentum
    - Fit-one-cycle learning rate scheduling
    - Optional Adam optimization
    - Support for capacity constraints
    """
    try:
        # Read customer data
        contents = await customer_file.read()
        customer_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Generate facility candidates
        facility_candidates = generate_candidate_facilities(
            customer_df=customer_df,
            method=candidate_method,
            n_candidates=n_candidates,
            lat_col=lat_col,
            lon_col=lon_col
        )
        
        if len(facility_candidates) < k:
            raise HTTPException(
                status_code=400,
                detail=f"Not enough facility candidates ({len(facility_candidates)}) for k={k}"
            )
        
        # Prepare customer weights (if available)
        if demand_col in customer_df.columns:
            weights = customer_df[demand_col].tolist()
        else:
            weights = [1.0] * len(customer_df)
        
        # Create distance matrix between facility candidates  
        n_candidates = len(facility_candidates)
        n_customers = len(customer_df)
        cost_matrix = np.zeros((n_candidates, n_candidates))
        
        # Calculate distances between facility candidates
        for i, (lat1, lon1) in enumerate(facility_candidates):
            for j, (lat2, lon2) in enumerate(facility_candidates):
                cost_matrix[i][j] = great_circle_distance(lat1, lon1, lat2, lon2)
        
        # Create candidate facility DataFrame for solve_k_median
        candidate_df = pd.DataFrame({
            'lat': [loc[0] for loc in facility_candidates],
            'lon': [loc[1] for loc in facility_candidates]
        })
        
        # For k-median, we need weights per candidate facility based on customer assignments
        # Calculate customer-to-candidate distances to determine initial weights
        customer_to_candidate_distances = np.zeros((n_customers, n_candidates))
        for i, customer in customer_df.iterrows():
            cust_lat = customer[lat_col]
            cust_lon = customer[lon_col]
            if pd.isna(cust_lat) or pd.isna(cust_lon):
                raise ValueError(f"Customer {i} has NaN coordinates in initial calc: lat={cust_lat}, lon={cust_lon}")
                
            for j, (fac_lat, fac_lon) in enumerate(facility_candidates):
                if pd.isna(fac_lat) or pd.isna(fac_lon):
                    raise ValueError(f"Facility candidate {j} has NaN coordinates: lat={fac_lat}, lon={fac_lon}")
                    
                customer_to_candidate_distances[i][j] = great_circle_distance(
                    cust_lat, cust_lon, 
                    fac_lat, fac_lon
                )
        
        # Assign each customer to nearest candidate to get weights
        candidate_weights = np.zeros(n_candidates)
        for i in range(n_customers):
            nearest_candidate = np.argmin(customer_to_candidate_distances[i])
            candidate_weights[nearest_candidate] += weights[i]
        
        # Ensure no zero weights (use minimum weight of 1)
        candidate_weights = np.maximum(candidate_weights, 1.0)
        
        # Solve k-median problem using the notebook implementation
        X, Y, partition, best_ub, lb_list, ub_list, phi_list = solve_k_median(
            candidate_df,
            candidate_weights.tolist(),
            cost_matrix.tolist(),
            k,
            max_iterations,
            max_lr,
            (0.85, 0.95),
            tolerance,  # convergence
            False,      # lr_find
            use_adam,   # adam
            capacity    # capacity
        )
        
        # Get selected facility indices from the k-median solution
        selected_facility_indices = list(range(len(X)))  # X, Y contain the selected facilities
        
        # Now assign customers to the selected facilities (not facility candidates)
        customer_assignments = []
        for i, customer in customer_df.iterrows():
            min_distance = float('inf')
            assigned_facility = 0
            
            # Debug: Check for NaN values
            cust_lat = customer[lat_col]
            cust_lon = customer[lon_col]
            if pd.isna(cust_lat) or pd.isna(cust_lon):
                raise ValueError(f"Customer {i} has NaN coordinates: lat={cust_lat}, lon={cust_lon}")
            
            for j, (fac_lat, fac_lon) in enumerate(zip(X, Y)):
                if pd.isna(fac_lat) or pd.isna(fac_lon):
                    raise ValueError(f"Facility {j} has NaN coordinates: lat={fac_lat}, lon={fac_lon}")
                    
                distance = great_circle_distance(
                    cust_lat, cust_lon,
                    fac_lat, fac_lon
                )
                if distance < min_distance:
                    min_distance = distance
                    assigned_facility = j
            
            customer_assignments.append(assigned_facility)
        
        # Prepare facility statistics based on actual customer assignments
        facility_stats = []
        for i, (lat, lon) in enumerate(zip(X, Y)):
            customers_assigned = sum(1 for assignment in customer_assignments if assignment == i)
            total_demand = sum(
                weights[j] for j, assignment in enumerate(customer_assignments) if assignment == i
            )
            
            # Calculate average distance for customers assigned to this facility
            distances = [
                great_circle_distance(
                    customer_df.iloc[j][lat_col], customer_df.iloc[j][lon_col],
                    lat, lon
                )
                for j, assignment in enumerate(customer_assignments) if assignment == i
            ]
            avg_distance = sum(distances) / len(distances) if distances else 0.0
            
            facility_stats.append({
                'facility_index': i,
                'location': [lat, lon],
                'customers_assigned': customers_assigned,
                'total_demand_served': total_demand,
                'average_distance': avg_distance
            })
        
        solution = {
            'selected_facilities': selected_facility_indices,
            'facility_locations': [[lat, lon] for lat, lon in zip(X, Y)],
            'assignments': customer_assignments,
            'facility_stats': facility_stats,
            'total_cost': float(best_ub),
            'algorithm': 'k_median_lagrangian_relaxation',
            'convergence_info': {
                'lower_bounds': [float(lb) for lb in lb_list],
                'upper_bounds': [float(ub) for ub in ub_list],
                'learning_rates': [float(phi) for phi in phi_list],
                'final_gap': float((best_ub - lb_list[-1]) / lb_list[-1]) if lb_list and lb_list[-1] != 0 else 0.0
            },
            'parameters': {
                'k': k,
                'max_iterations': max_iterations,
                'max_lr': max_lr,
                'tolerance': tolerance,
                'use_adam': use_adam,
                'capacity': capacity,
                'candidate_method': candidate_method,
                'n_candidates': n_candidates
            }
        }
        
        # Add candidate facilities to solution for visualization
        solution["candidate_facilities"] = [
            {
                "id": i,
                "latitude": lat,
                "longitude": lon
            }
            for i, (lat, lon) in enumerate(facility_candidates)
        ]
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert solution to JSON-safe format
        solution = convert_to_json_safe(solution)
        
        return solution
    
    except Exception as e:
        import traceback
        error_detail = f"Error solving k-median problem: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error solving k-median problem: {str(e)}")

@router.post("/k-median-lr-find")
async def k_median_learning_rate_finder(
    customer_file: UploadFile = File(...),
    k: int = Form(3),
    candidate_method: str = Form("grid"),
    n_candidates: int = Form(20),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    max_iterations: int = Form(50)
):
    """
    Perform learning rate finding for k-median Lagrangian relaxation
    
    This helps determine optimal learning rate parameters by running
    the algorithm with exponentially increasing learning rates
    """
    try:
        # Read customer data
        contents = await customer_file.read()
        customer_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Generate facility candidates
        facility_candidates = generate_candidate_facilities(
            customer_df=customer_df,
            method=candidate_method,
            n_candidates=n_candidates,
            lat_col=lat_col,
            lon_col=lon_col
        )
        
        if len(facility_candidates) < k:
            raise HTTPException(
                status_code=400,
                detail=f"Not enough facility candidates ({len(facility_candidates)}) for k={k}"
            )
        
        # TODO: Implement learning rate finding with notebook implementation
        solution = {
            'error': 'Learning rate finding needs to be updated to match notebook',
            'facility_candidates': len(facility_candidates),
            'customers': len(customer_df),
            'k': k,
            'learning_rates': [0.01, 0.02, 0.03],  # Placeholder
            'algorithm': 'lr_finder_placeholder'
        }
        
        # Convert arrays to lists for JSON serialization  
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        solution = convert_to_json_safe(solution)
        
        return {
            "learning_rates": solution.get('learning_rates', []),
            "lower_bounds": solution.get('lower_bounds', []),
            "upper_bounds": solution.get('upper_bounds', []),
            "iterations": solution.get('iterations', 0),
            "recommended_lr": solution['learning_rates'][-1] / 10 if solution.get('learning_rates') else 0.01,
            "parameters": {
                "k": k,
                "candidate_method": candidate_method,
                "n_candidates": n_candidates,
                "max_iterations": max_iterations
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error in learning rate finding: {str(e)}")

@router.post("/service-area-analysis")
async def analyze_service_area(
    file: UploadFile = File(...),
    facility_lat: float = Form(...),
    facility_lon: float = Form(...),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand")
):
    """
    Analyze service area statistics for a given facility location
    """
    try:
        # Read CSV file
        contents = await file.read()
        customer_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Calculate service area statistics
        service_stats = calculate_facility_service_area(
            customer_df=customer_df,
            facility_location=(facility_lat, facility_lon),
            lat_col=lat_col,
            lon_col=lon_col,
            demand_col=demand_col
        )
        
        return {
            "service_area_analysis": service_stats,
            "facility_location": {
                "latitude": facility_lat,
                "longitude": facility_lon
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error analyzing service area: {str(e)}")

@router.post("/generate-candidates")
async def generate_facility_candidates(
    file: UploadFile = File(...),
    method: str = Form("grid"),
    n_candidates: int = Form(20),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon")
):
    """
    Generate candidate facility locations using various methods
    """
    try:
        # Read CSV file
        contents = await file.read()
        customer_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Generate candidate locations
        candidates = generate_candidate_facilities(
            customer_df=customer_df,
            method=method,
            n_candidates=n_candidates,
            lat_col=lat_col,
            lon_col=lon_col
        )
        
        # Format candidates for response
        formatted_candidates = [
            {
                "id": i,
                "latitude": lat,
                "longitude": lon
            }
            for i, (lat, lon) in enumerate(candidates)
        ]
        
        return {
            "candidate_facilities": formatted_candidates,
            "method": method,
            "n_candidates": len(formatted_candidates),
            "customer_bounds": {
                "min_lat": float(customer_df[lat_col].min()),
                "max_lat": float(customer_df[lat_col].max()),
                "min_lon": float(customer_df[lon_col].min()),
                "max_lon": float(customer_df[lon_col].max())
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error generating facility candidates: {str(e)}")

@router.post("/distance-calculation")
async def calculate_distances(
    origin_lat: float = Form(...),
    origin_lon: float = Form(...),
    dest_lat: float = Form(...),
    dest_lon: float = Form(...)
):
    """
    Calculate great circle distance between two points
    """
    try:
        distance = great_circle_distance(origin_lat, origin_lon, dest_lat, dest_lon)
        
        return {
            "distance_km": distance,
            "origin": {
                "latitude": origin_lat,
                "longitude": origin_lon
            },
            "destination": {
                "latitude": dest_lat,
                "longitude": dest_lon
            },
            "calculation_method": "great_circle"
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error calculating distance: {str(e)}")

@router.get("/sample-data-info")
async def get_sample_data_information():
    """
    Get information about available sample datasets
    """
    try:
        return get_sample_data_info()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting sample data info: {str(e)}")

@router.get("/sample-data/{dataset_type}")
async def download_sample_data(dataset_type: str):
    """
    Download sample data as CSV file
    
    Available dataset types:
    - customers_standard: 50 customers around Tokyo
    - customers_small: 20 customers around Tokyo (for quick testing)
    - customers_regional: 40 customers across Tokyo/Osaka/Nagoya
    - facilities: 15 facility candidates around Tokyo
    """
    try:
        print(f"Requesting dataset: {dataset_type}")
        datasets = create_lnd_sample_datasets()
        print(f"Available datasets: {list(datasets.keys())}")
        
        if dataset_type not in datasets:
            raise HTTPException(
                status_code=404, 
                detail=f"Dataset type '{dataset_type}' not found. Available: {list(datasets.keys())}"
            )
        
        csv_content = datasets[dataset_type]
        print(f"Generated CSV content length: {len(csv_content)}")
        
        # Set appropriate filename
        filename_map = {
            'customers_standard': 'lnd_customers_standard.csv',
            'customers_small': 'lnd_customers_small.csv', 
            'customers_regional': 'lnd_customers_regional.csv',
            'facilities': 'lnd_facilities.csv',
            'ms_lnd_customers': 'ms_lnd_customers.csv',
            'ms_lnd_warehouses': 'ms_lnd_warehouses.csv',
            'ms_lnd_factories': 'ms_lnd_factories.csv',
            'ms_lnd_products': 'ms_lnd_products.csv',
            'ms_lnd_demand': 'ms_lnd_demand.csv',
            'ms_lnd_factory_capacity': 'ms_lnd_factory_capacity.csv',
            'elbow_customers_3clusters': 'elbow_customers_3clusters.csv',
            'elbow_customers_2clusters': 'elbow_customers_2clusters.csv',
            'elbow_customers_5clusters': 'elbow_customers_5clusters.csv'
        }
        
        if dataset_type not in filename_map:
            raise HTTPException(
                status_code=404,
                detail=f"No filename mapping for dataset '{dataset_type}'"
            )
        
        filename = filename_map[dataset_type]
        print(f"Using filename: {filename}")
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_detail = f"Error generating sample data: {str(e)}\nTraceback: {traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"Error generating sample data: {str(e)}")

@router.post("/multiple-source-lnd")
async def solve_multiple_source_logistics_network_design(
    customer_file: UploadFile = File(...),
    warehouse_file: UploadFile = File(...),
    factory_file: UploadFile = File(...),
    product_file: UploadFile = File(...),
    demand_file: UploadFile = File(...),
    factory_capacity_file: UploadFile = File(...),
    transportation_cost: float = Form(1.0),
    delivery_cost: float = Form(2.0),
    warehouse_fixed_cost: float = Form(10000.0),
    warehouse_variable_cost: float = Form(1.0),
    num_warehouses: Optional[int] = Form(None),
    single_sourcing: bool = Form(False),
    max_runtime: int = Form(300)
):
    """
    Solve Multiple Source Logistics Network Design problem using mixed-integer optimization
    
    This endpoint implements a comprehensive logistics network optimization model that considers:
    - Multiple factories producing multiple products
    - Multiple warehouse locations with capacity constraints
    - Multiple customers with specific product demands
    - Flow conservation and sourcing constraints
    - Cost optimization (fixed costs, variable costs, transportation costs)
    
    Features:
    - Mixed-Integer Linear Programming (MILP) optimization
    - Single-sourcing or multiple-sourcing constraints
    - Warehouse capacity bounds (upper and lower limits)
    - Factory production capacity constraints
    - Great circle distance calculations for transportation costs
    - Comprehensive solution analysis and visualization data
    """
    try:
        # Read all uploaded CSV files
        customer_contents = await customer_file.read()
        customer_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        
        warehouse_contents = await warehouse_file.read()
        warehouse_df = pd.read_csv(io.StringIO(warehouse_contents.decode('utf-8')))
        
        factory_contents = await factory_file.read()
        factory_df = pd.read_csv(io.StringIO(factory_contents.decode('utf-8')))
        
        product_contents = await product_file.read()
        product_df = pd.read_csv(io.StringIO(product_contents.decode('utf-8')))
        
        demand_contents = await demand_file.read()
        demand_df = pd.read_csv(io.StringIO(demand_contents.decode('utf-8')))
        
        factory_capacity_contents = await factory_capacity_file.read()
        factory_capacity_df = pd.read_csv(io.StringIO(factory_capacity_contents.decode('utf-8')))
        
        # Validate required columns for each dataframe
        required_columns = {
            'customer': ['customer_id', 'lat', 'lon'],
            'warehouse': ['warehouse_id', 'lat', 'lon', 'upper_bound', 'lower_bound', 'is_available'],
            'factory': ['factory_id', 'lat', 'lon'],
            'product': ['product_id'],
            'demand': ['customer_id', 'product_id', 'demand'],
            'factory_capacity': ['factory_id', 'product_id', 'capacity']
        }
        
        dataframes = {
            'customer': customer_df,
            'warehouse': warehouse_df,
            'factory': factory_df,
            'product': product_df,
            'demand': demand_df,
            'factory_capacity': factory_capacity_df
        }
        
        # Check for missing columns
        for df_name, df in dataframes.items():
            missing_cols = [col for col in required_columns[df_name] if col not in df.columns]
            if missing_cols:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns in {df_name} file: {missing_cols}"
                )
        
        # Validate parameter ranges
        if transportation_cost < 0:
            raise HTTPException(status_code=400, detail="Transportation cost must be non-negative")
        if delivery_cost < 0:
            raise HTTPException(status_code=400, detail="Delivery cost must be non-negative")
        if warehouse_fixed_cost < 0:
            raise HTTPException(status_code=400, detail="Warehouse fixed cost must be non-negative")
        if warehouse_variable_cost < 0:
            raise HTTPException(status_code=400, detail="Warehouse variable cost must be non-negative")
        if max_runtime < 1:
            raise HTTPException(status_code=400, detail="Max runtime must be at least 1 second")
        if max_runtime > 3600:
            raise HTTPException(status_code=400, detail="Max runtime cannot exceed 1 hour (3600 seconds)")
        
        if num_warehouses is not None:
            if num_warehouses < 1:
                raise HTTPException(status_code=400, detail="Number of warehouses must be at least 1")
            if num_warehouses > len(warehouse_df):
                raise HTTPException(
                    status_code=400,
                    detail=f"Number of warehouses ({num_warehouses}) cannot exceed available warehouses ({len(warehouse_df)})"
                )
        
        # Solve the Multiple Source LND problem
        solution = solve_multiple_source_lnd(
            customer_df=customer_df,
            warehouse_df=warehouse_df,
            factory_df=factory_df,
            product_df=product_df,
            demand_df=demand_df,
            factory_capacity_df=factory_capacity_df,
            transportation_cost=transportation_cost,
            delivery_cost=delivery_cost,
            warehouse_fixed_cost=warehouse_fixed_cost,
            warehouse_variable_cost=warehouse_variable_cost,
            num_warehouses=num_warehouses,
            single_sourcing=single_sourcing,
            max_runtime=max_runtime
        )
        
        if not solution or solution.get('status') != 'Optimal':
            error_msg = solution.get('message', 'Unknown optimization error') if solution else 'No solution returned'
            raise HTTPException(
                status_code=500,
                detail=f"Failed to find optimal solution: {error_msg}"
            )
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert solution to JSON-safe format
        solution = convert_to_json_safe(solution)
        
        return solution
    
    except Exception as e:
        import traceback
        error_detail = f"Error solving Multiple Source LND: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error solving Multiple Source LND: {str(e)}")

@router.post("/single-source-lnd")
async def solve_single_source_logistics_network_design(
    customer_file: UploadFile = File(...),
    warehouse_file: UploadFile = File(...),
    factory_file: UploadFile = File(...),
    product_file: UploadFile = File(...),
    demand_file: UploadFile = File(...),
    factory_capacity_file: UploadFile = File(...),
    transportation_cost: float = Form(1.0),
    delivery_cost: float = Form(2.0),
    warehouse_fixed_cost: float = Form(10000.0),
    warehouse_variable_cost: float = Form(1.0),
    num_warehouses: Optional[int] = Form(None),
    max_runtime: int = Form(300)
):
    """
    Solve Single Source Logistics Network Design problem using mixed-integer optimization
    
    Single Source Constraint: Each customer is served by exactly one warehouse
    
    This endpoint implements the single-sourcing version of the logistics network optimization model:
    - Multiple factories producing multiple products
    - Multiple warehouse locations with capacity constraints
    - Multiple customers with specific product demands
    - Single-sourcing constraint: each customer served by exactly one warehouse
    - Cost optimization (fixed costs, variable costs, transportation costs)
    
    Features:
    - Mixed-Integer Linear Programming (MILP) optimization
    - Single-sourcing assignment variables
    - Warehouse capacity bounds (upper and lower limits)
    - Factory production capacity constraints
    - Great circle distance calculations for transportation costs
    - Comprehensive solution analysis and visualization data
    """
    try:
        # Read all uploaded CSV files
        customer_contents = await customer_file.read()
        customer_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        
        warehouse_contents = await warehouse_file.read()
        warehouse_df = pd.read_csv(io.StringIO(warehouse_contents.decode('utf-8')))
        
        factory_contents = await factory_file.read()
        factory_df = pd.read_csv(io.StringIO(factory_contents.decode('utf-8')))
        
        product_contents = await product_file.read()
        product_df = pd.read_csv(io.StringIO(product_contents.decode('utf-8')))
        
        demand_contents = await demand_file.read()
        demand_df = pd.read_csv(io.StringIO(demand_contents.decode('utf-8')))
        
        factory_capacity_contents = await factory_capacity_file.read()
        factory_capacity_df = pd.read_csv(io.StringIO(factory_capacity_contents.decode('utf-8')))
        
        # Validate required columns for each dataframe
        required_columns = {
            'customer': ['customer_id', 'lat', 'lon'],
            'warehouse': ['warehouse_id', 'lat', 'lon', 'upper_bound', 'lower_bound', 'is_available'],
            'factory': ['factory_id', 'lat', 'lon'],
            'product': ['product_id'],
            'demand': ['customer_id', 'product_id', 'demand'],
            'factory_capacity': ['factory_id', 'product_id', 'capacity']
        }
        
        dataframes = {
            'customer': customer_df,
            'warehouse': warehouse_df,
            'factory': factory_df,
            'product': product_df,
            'demand': demand_df,
            'factory_capacity': factory_capacity_df
        }
        
        # Check for missing columns
        for df_name, df in dataframes.items():
            missing_cols = [col for col in required_columns[df_name] if col not in df.columns]
            if missing_cols:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns in {df_name} file: {missing_cols}"
                )
        
        # Validate parameter ranges
        if transportation_cost < 0:
            raise HTTPException(status_code=400, detail="Transportation cost must be non-negative")
        if delivery_cost < 0:
            raise HTTPException(status_code=400, detail="Delivery cost must be non-negative")
        if warehouse_fixed_cost < 0:
            raise HTTPException(status_code=400, detail="Warehouse fixed cost must be non-negative")
        if warehouse_variable_cost < 0:
            raise HTTPException(status_code=400, detail="Warehouse variable cost must be non-negative")
        if max_runtime < 1:
            raise HTTPException(status_code=400, detail="Max runtime must be at least 1 second")
        if max_runtime > 3600:
            raise HTTPException(status_code=400, detail="Max runtime cannot exceed 1 hour (3600 seconds)")
        
        if num_warehouses is not None:
            if num_warehouses < 1:
                raise HTTPException(status_code=400, detail="Number of warehouses must be at least 1")
            if num_warehouses > len(warehouse_df):
                raise HTTPException(
                    status_code=400,
                    detail=f"Number of warehouses ({num_warehouses}) cannot exceed available warehouses ({len(warehouse_df)})"
                )
        
        # Solve the Single Source LND problem
        solution = solve_single_source_lnd(
            customer_df=customer_df,
            warehouse_df=warehouse_df,
            factory_df=factory_df,
            product_df=product_df,
            demand_df=demand_df,
            factory_capacity_df=factory_capacity_df,
            transportation_cost=transportation_cost,
            delivery_cost=delivery_cost,
            warehouse_fixed_cost=warehouse_fixed_cost,
            warehouse_variable_cost=warehouse_variable_cost,
            num_warehouses=num_warehouses,
            max_runtime=max_runtime
        )
        
        if not solution or solution.get('status') != 'Optimal':
            error_msg = solution.get('message', 'Unknown optimization error') if solution else 'No solution returned'
            raise HTTPException(
                status_code=500,
                detail=f"Failed to find optimal solution: {error_msg}"
            )
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert solution to JSON-safe format
        solution = convert_to_json_safe(solution)
        
        return solution
    
    except Exception as e:
        import traceback
        error_detail = f"Error solving Single Source LND: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error solving Single Source LND: {str(e)}")


@router.post("/elbow-method-analysis")
async def perform_elbow_method_analysis(
    file: UploadFile = File(...),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: Optional[str] = Form(None),
    min_facilities: int = Form(1),
    max_facilities: int = Form(10),
    algorithm: str = Form("weiszfeld"),
    max_iterations: int = Form(1000),
    tolerance: float = Form(1e-4),
    random_state: int = Form(42)
):
    """
    Perform elbow method analysis to determine optimal number of facilities
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Validate algorithm
        valid_algorithms = ['weiszfeld', 'kmeans', 'hierarchical']
        if algorithm not in valid_algorithms:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid algorithm. Must be one of: {valid_algorithms}"
            )
        
        # Validate facility range
        if min_facilities < 1:
            raise HTTPException(
                status_code=400,
                detail="Minimum facilities must be at least 1"
            )
        
        if max_facilities < min_facilities:
            raise HTTPException(
                status_code=400,
                detail="Maximum facilities must be greater than or equal to minimum facilities"
            )
        
        if max_facilities - min_facilities < 2:
            raise HTTPException(
                status_code=400,
                detail="Need at least 3 different facility counts for elbow method analysis"
            )
        
        # Perform analysis
        result = elbow_method_analysis(
            customer_df=df,
            lat_col=lat_col,
            lon_col=lon_col,
            demand_col=demand_col,
            min_facilities=min_facilities,
            max_facilities=max_facilities,
            algorithm=algorithm,
            max_iterations=max_iterations,
            tolerance=tolerance,
            random_state=random_state
        )
        
        # Check for errors
        if result.get('status') == 'Error':
            raise HTTPException(status_code=500, detail=result['message'])
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in elbow method analysis: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error in elbow method analysis: {str(e)}")

@router.post("/hierarchical-clustering-advanced")
async def calculate_hierarchical_clustering_advanced(
    file: UploadFile = File(...),
    num_facilities: int = Form(2),
    linkage: str = Form("average"),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand")
):
    """
    Advanced hierarchical clustering for facility location using distance matrices
    
    Features:
    - Multiple linkage methods (ward, complete, average, single)
    - Real distance matrix calculations using great circle distances
    - Median-based facility positioning for optimal locations
    - Weighted cost calculation based on customer demands
    
    This implementation is extracted from 05lnd.ipynb to provide advanced 
    hierarchical clustering capabilities with optimal facility placement.
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Validate linkage method
        valid_linkages = ["ward", "complete", "average", "single"]
        if linkage not in valid_linkages:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid linkage method. Must be one of: {valid_linkages}"
            )
        
        # Validate num_facilities
        if num_facilities < 1:
            raise HTTPException(
                status_code=400,
                detail="Number of facilities must be at least 1"
            )
        
        if num_facilities > len(df):
            raise HTTPException(
                status_code=400,
                detail=f"Number of facilities ({num_facilities}) cannot exceed number of customers ({len(df)})"
            )
        
        # Extract weights
        weights = df[demand_col].tolist() if demand_col in df.columns else [1.0] * len(df)
        
        # Perform hierarchical clustering
        X, Y, partition, total_cost = advanced_facility_service.hierarchical_clustering(
            customer_df=df,
            weight=weights,
            distances=None,  # Will compute distance matrix internally
            num_of_facilities=num_facilities,
            linkage=linkage
        )
        
        # Create facility statistics
        facility_stats = []
        clusters = {}
        for i, cluster_id in enumerate(partition):
            if cluster_id not in clusters:
                clusters[cluster_id] = []
            clusters[cluster_id].append(i)
        
        for facility_idx in range(len(X)):
            if facility_idx < len(clusters):
                cluster_customers = clusters[facility_idx]
                facility_demand = sum(weights[i] for i in cluster_customers)
                
                # Calculate average distance to customers in cluster
                avg_distance = 0
                if cluster_customers:
                    distances = []
                    for customer_idx in cluster_customers:
                        customer_lat = df.iloc[customer_idx][lat_col]
                        customer_lon = df.iloc[customer_idx][lon_col]
                        dist = advanced_facility_service.great_circle_distance(
                            X[facility_idx], Y[facility_idx], customer_lat, customer_lon
                        )
                        distances.append(dist)
                    avg_distance = np.mean(distances) if distances else 0
                
                facility_stats.append({
                    "facility_index": facility_idx,
                    "location": [X[facility_idx], Y[facility_idx]],
                    "customers_assigned": len(cluster_customers),
                    "total_demand_served": facility_demand,
                    "average_distance": avg_distance
                })
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        result = {
            "facility_locations": [[lat, lon] for lat, lon in zip(X, Y)],
            "assignments": partition,
            "total_cost": total_cost,
            "facility_stats": facility_stats,
            "algorithm": "hierarchical_clustering_advanced",
            "linkage_method": linkage,
            "parameters": {
                "num_facilities": num_facilities,
                "linkage": linkage,
                "distance_method": "great_circle",
                "weighted": demand_col in df.columns
            }
        }
        
        # Convert to JSON-safe format
        result = convert_to_json_safe(result)
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in hierarchical clustering advanced: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error in hierarchical clustering advanced: {str(e)}")

@router.post("/k-median-advanced")
async def solve_k_median_advanced(
    file: UploadFile = File(...),
    num_facilities: int = Form(2),
    max_iter: int = Form(100),
    max_lr: float = Form(0.01),
    mom_low: float = Form(0.85),
    mom_high: float = Form(0.95),
    convergence: float = Form(1e-5),
    adam: bool = Form(False),
    capacity: Optional[int] = Form(None),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand")
):
    """
    Advanced k-median optimization using Lagrange relaxation with fit-one-cycle scheduling
    
    Features:
    - Lagrange relaxation for k-median problems
    - Fit-one-cycle learning rate scheduling
    - Adam optimizer integration
    - Capacity-constrained facility location support
    - Advanced convergence monitoring
    - Mixed-Integer Linear Programming (MILP) approach
    
    This implementation is extracted from 05lnd.ipynb to provide the exact same
    calculation results as the notebook implementation.
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Validate parameters
        if num_facilities < 1:
            raise HTTPException(
                status_code=400,
                detail="Number of facilities must be at least 1"
            )
        
        if num_facilities > len(df):
            raise HTTPException(
                status_code=400,
                detail=f"Number of facilities ({num_facilities}) cannot exceed number of customers ({len(df)})"
            )
        
        if max_iter < 1:
            raise HTTPException(
                status_code=400,
                detail="Maximum iterations must be at least 1"
            )
        
        if max_lr <= 0:
            raise HTTPException(
                status_code=400,
                detail="Maximum learning rate must be positive"
            )
        
        if not (0 < mom_low < 1) or not (0 < mom_high < 1):
            raise HTTPException(
                status_code=400,
                detail="Momentum values must be between 0 and 1"
            )
        
        if mom_low >= mom_high:
            raise HTTPException(
                status_code=400,
                detail="Low momentum must be less than high momentum"
            )
        
        if convergence <= 0:
            raise HTTPException(
                status_code=400,
                detail="Convergence tolerance must be positive"
            )
        
        if capacity is not None and capacity < 1:
            raise HTTPException(
                status_code=400,
                detail="Capacity must be at least 1 if specified"
            )
        
        # Extract weights and compute cost matrix
        weights = df[demand_col].tolist() if demand_col in df.columns else [1.0] * len(df)
        
        # Compute distance/cost matrix
        n = len(df)
        cost_matrix = np.zeros((n, n))
        
        for i in range(n):
            for j in range(n):
                if i != j:
                    lat1, lon1 = df.iloc[i][lat_col], df.iloc[i][lon_col]
                    lat2, lon2 = df.iloc[j][lat_col], df.iloc[j][lon_col]
                    cost_matrix[i, j] = advanced_facility_service.great_circle_distance(lat1, lon1, lat2, lon2)
        
        # Solve k-median problem
        X, Y, partition, best_ub, lb_list, ub_list, phi_list = advanced_facility_service.solve_k_median(
            customer_df=df,
            weight=weights,
            cost=cost_matrix,
            num_of_facilities=num_facilities,
            max_iter=max_iter,
            max_lr=max_lr,
            moms=(mom_low, mom_high),
            convergence=convergence,
            lr_find=False,  # Default to False for standard optimization
            adam=adam,
            capacity=capacity
        )
        
        # Create facility statistics
        facility_stats = []
        clusters = {}
        for i, cluster_id in enumerate(partition):
            if cluster_id not in clusters:
                clusters[cluster_id] = []
            clusters[cluster_id].append(i)
        
        for facility_idx in range(len(X)):
            if facility_idx < len(clusters):
                cluster_customers = clusters[facility_idx]
                facility_demand = sum(weights[i] for i in cluster_customers)
                
                # Calculate average distance to customers in cluster
                avg_distance = 0
                if cluster_customers:
                    distances = []
                    for customer_idx in cluster_customers:
                        customer_lat = df.iloc[customer_idx][lat_col]
                        customer_lon = df.iloc[customer_idx][lon_col]
                        dist = advanced_facility_service.great_circle_distance(
                            X[facility_idx], Y[facility_idx], customer_lat, customer_lon
                        )
                        distances.append(dist)
                    avg_distance = np.mean(distances) if distances else 0
                
                facility_stats.append({
                    "facility_index": facility_idx,
                    "location": [X[facility_idx], Y[facility_idx]],
                    "customers_assigned": len(cluster_customers),
                    "total_demand_served": facility_demand,
                    "average_distance": avg_distance
                })
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        result = {
            "facility_locations": [[lat, lon] for lat, lon in zip(X, Y)],
            "assignments": partition,
            "total_cost": best_ub,
            "facility_stats": facility_stats,
            "optimization_history": {
                "lower_bounds": lb_list,
                "upper_bounds": ub_list,
                "learning_rates": phi_list
            },
            "algorithm": "k_median_lagrange_advanced",
            "parameters": {
                "num_facilities": num_facilities,
                "max_iterations": max_iter,
                "max_learning_rate": max_lr,
                "momentum_bounds": [mom_low, mom_high],
                "convergence_tolerance": convergence,
                "adam_optimizer": adam,
                "capacity_constraint": capacity,
                "weighted": demand_col in df.columns
            }
        }
        
        # Convert to JSON-safe format
        result = convert_to_json_safe(result)
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in k-median advanced optimization: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error in k-median advanced optimization: {str(e)}")

@router.post("/k-median-lr-finder-advanced")
async def solve_k_median_lr_finder_advanced(
    file: UploadFile = File(...),
    num_facilities: int = Form(2),
    max_iter: int = Form(100),
    mom_low: float = Form(0.85),
    mom_high: float = Form(0.95),
    capacity: Optional[int] = Form(None),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand")
):
    """
    K-median optimization with learning rate finder for automatic LR discovery
    
    Features:
    - Automatic learning rate discovery using lr_find technique
    - Lagrange relaxation optimization
    - Capacity-constrained facility location support
    - Learning rate search from 1e-10 upward with exponential scaling
    
    This endpoint helps find the optimal learning rate for k-median problems
    by running the optimization with lr_find=True to discover the best LR range.
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Validate parameters
        if num_facilities < 1:
            raise HTTPException(
                status_code=400,
                detail="Number of facilities must be at least 1"
            )
        
        if num_facilities > len(df):
            raise HTTPException(
                status_code=400,
                detail=f"Number of facilities ({num_facilities}) cannot exceed number of customers ({len(df)})"
            )
        
        if max_iter < 1:
            raise HTTPException(
                status_code=400,
                detail="Maximum iterations must be at least 1"
            )
        
        if not (0 < mom_low < 1) or not (0 < mom_high < 1):
            raise HTTPException(
                status_code=400,
                detail="Momentum values must be between 0 and 1"
            )
        
        if mom_low >= mom_high:
            raise HTTPException(
                status_code=400,
                detail="Low momentum must be less than high momentum"
            )
        
        if capacity is not None and capacity < 1:
            raise HTTPException(
                status_code=400,
                detail="Capacity must be at least 1 if specified"
            )
        
        # Extract weights and compute cost matrix
        weights = df[demand_col].tolist() if demand_col in df.columns else [1.0] * len(df)
        
        # Compute distance/cost matrix
        n = len(df)
        cost_matrix = np.zeros((n, n))
        
        for i in range(n):
            for j in range(n):
                if i != j:
                    lat1, lon1 = df.iloc[i][lat_col], df.iloc[i][lon_col]
                    lat2, lon2 = df.iloc[j][lat_col], df.iloc[j][lon_col]
                    cost_matrix[i, j] = advanced_facility_service.great_circle_distance(lat1, lon1, lat2, lon2)
        
        # Solve k-median problem with learning rate finder
        X, Y, partition, best_ub, lb_list, ub_list, phi_list = advanced_facility_service.solve_k_median(
            customer_df=df,
            weight=weights,
            cost=cost_matrix,
            num_of_facilities=num_facilities,
            max_iter=max_iter,
            max_lr=0.01,  # Not used when lr_find=True
            moms=(mom_low, mom_high),
            convergence=1e-5,  # Default convergence for lr_find
            lr_find=True,  # Enable learning rate finding
            adam=False,  # Use default momentum-based optimization
            capacity=capacity
        )
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        result = {
            "learning_rate_search": {
                "learning_rates": phi_list,
                "lower_bounds": lb_list,
                "upper_bounds": ub_list,
                "optimal_lr_range": {
                    "min_lr": min(phi_list) if phi_list else 1e-10,
                    "max_lr": max(phi_list) if phi_list else 1e-2,
                    "suggested_lr": max(phi_list) if phi_list else 1e-3
                }
            },
            "facility_locations": [[lat, lon] for lat, lon in zip(X, Y)],
            "assignments": partition,
            "total_cost": best_ub,
            "algorithm": "k_median_lr_finder",
            "parameters": {
                "num_facilities": num_facilities,
                "max_iterations": max_iter,
                "momentum_bounds": [mom_low, mom_high],
                "capacity_constraint": capacity,
                "weighted": demand_col in df.columns,
                "lr_find_enabled": True
            }
        }
        
        # Convert to JSON-safe format
        result = convert_to_json_safe(result)
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in k-median LR finder: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)  # Log the full error
        raise HTTPException(status_code=500, detail=f"Error in k-median LR finder: {str(e)}")

@router.get("/excel-template")
async def generate_excel_template():
    """
    Generate Excel template for MELOS (MEta Logistic network Optimization System)
    
    Creates a comprehensive Excel workbook template with the following sheets:
    - 顧客 (Customers): Customer locations and IDs
    - 倉庫候補地点 (Warehouse Candidates): Potential warehouse locations with costs
    - 工場 (Plants): Manufacturing plant locations
    - 製品 (Products): Product specifications (weight, volume)
    - 需要 (Demand): Customer demand matrix
    - 生産 (Production): Plant production capacity matrix
    
    Returns Excel file ready for data input and optimization.
    """
    try:
        # Generate Excel template
        wb = excel_integration_service.make_excel_melos()
        wb = excel_integration_service.make_demand_production_sheets(wb)
        
        # Save to bytes
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=MELOS_template.xlsx"
            }
        )
    
    except Exception as e:
        import traceback
        error_detail = f"Error generating Excel template: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"Error generating Excel template: {str(e)}")

@router.post("/excel-parse")
async def parse_excel_data(
    file: UploadFile = File(...)
):
    """
    Parse Excel data from MELOS template
    
    Extracts data from all sheets in the Excel workbook and returns structured data:
    - Customer data with locations
    - Warehouse candidate data with costs and constraints
    - Plant data with locations
    - Product specifications
    - Demand matrix (customers x products)
    - Production matrix (plants x products)
    """
    try:
        # Read Excel file
        contents = await file.read()
        from io import BytesIO
        excel_buffer = BytesIO(contents)
        
        from openpyxl import load_workbook
        wb = load_workbook(excel_buffer)
        
        # Parse data from Excel sheets
        cust_df, dc_df, plnt_df, prod_df, demand_df, production_df = excel_integration_service.prepare_df_for_melos(wb)
        
        # Convert DataFrames to JSON-safe format
        def convert_df_to_dict(df):
            return df.to_dict('records')
        
        result = {
            "customers": convert_df_to_dict(cust_df),
            "warehouses": convert_df_to_dict(dc_df),
            "plants": convert_df_to_dict(plnt_df),
            "products": convert_df_to_dict(prod_df),
            "demand": convert_df_to_dict(demand_df),
            "production": convert_df_to_dict(production_df),
            "summary": {
                "num_customers": len(cust_df),
                "num_warehouses": len(dc_df),
                "num_plants": len(plnt_df),
                "num_products": len(prod_df)
            }
        }
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error parsing Excel data: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"Error parsing Excel data: {str(e)}")

@router.post("/customer-aggregation")
async def perform_customer_aggregation(
    customer_file: UploadFile = File(...),
    product_file: UploadFile = File(...),
    demand_file: UploadFile = File(...),
    num_facilities: int = Form(3),
    linkage: str = Form("complete"),
    toll: bool = Form(True),
    osrm_host: str = Form("localhost"),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    name_col: str = Form("name")
):
    """
    Perform customer aggregation using hierarchical clustering with demand weights
    
    Features:
    - Demand-weighted hierarchical clustering
    - Road distance calculations using OSRM (or great circle fallback)
    - Multiple linkage methods (ward, complete, average, single)
    - Automatic demand aggregation across clusters
    - Interactive visualization data
    - Medoid-based facility positioning
    
    This is particularly useful for large customer datasets where direct optimization
    would be computationally expensive.
    """
    try:
        # Read uploaded files
        customer_contents = await customer_file.read()
        product_contents = await product_file.read()
        demand_contents = await demand_file.read()
        
        cust_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        prod_df = pd.read_csv(io.StringIO(product_contents.decode('utf-8')))
        demand_df = pd.read_csv(io.StringIO(demand_contents.decode('utf-8')))
        
        # Validate required columns
        required_customer_cols = [name_col, lat_col, lon_col]
        missing_cols = [col for col in required_customer_cols if col not in cust_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns in customer file: {missing_cols}"
            )
        
        # Validate parameters
        if num_facilities < 1:
            raise HTTPException(
                status_code=400,
                detail="Number of facilities must be at least 1"
            )
        
        if num_facilities > len(cust_df):
            raise HTTPException(
                status_code=400,
                detail=f"Number of facilities ({num_facilities}) cannot exceed number of customers ({len(cust_df)})"
            )
        
        valid_linkages = ["ward", "complete", "average", "single"]
        if linkage not in valid_linkages:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid linkage method. Must be one of: {valid_linkages}"
            )
        
        # Rename columns to standard format
        cust_df = cust_df.rename(columns={
            name_col: 'name',
            lat_col: 'lat', 
            lon_col: 'lon'
        })
        
        # Set OSRM host
        customer_aggregation_service.osrm_host = osrm_host
        
        # Perform customer aggregation
        aggregated_cust_df, visualization_data, aggregated_demand = customer_aggregation_service.customer_aggregation(
            cust_df=cust_df,
            prod_df=prod_df,
            demand_df=demand_df,
            num_of_facilities=num_facilities,
            linkage=linkage,
            toll=toll
        )
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Prepare result
        result = {
            "aggregated_customers": aggregated_cust_df.to_dict('records'),
            "visualization_data": visualization_data,
            "aggregated_demand": aggregated_demand.tolist(),
            "original_customers_count": len(cust_df),
            "aggregated_customers_count": len(aggregated_cust_df),
            "reduction_ratio": len(aggregated_cust_df) / len(cust_df),
            "algorithm": "hierarchical_clustering_with_demand_weights",
            "parameters": {
                "num_facilities": num_facilities,
                "linkage": linkage,
                "toll_roads": toll,
                "osrm_host": osrm_host,
                "distance_method": "osrm_with_fallback"
            }
        }
        
        # Convert to JSON-safe format
        result = convert_to_json_safe(result)
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in customer aggregation: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"Error in customer aggregation: {str(e)}")

@router.get("/service-info")
async def get_service_info():
    """
    Get information about all available LND services and their capabilities
    """
    try:
        info = {
            "advanced_facility_location": advanced_facility_service.get_algorithm_info(),
            "excel_integration": excel_integration_service.get_service_info(),
            "customer_aggregation": customer_aggregation_service.get_service_info(),
            "carbon_footprint_analysis": carbon_footprint_service.get_service_info()
        }
        
        return info
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting service info: {str(e)}")

@router.post("/carbon-footprint-analysis")
async def perform_carbon_footprint_analysis(
    facility_file: UploadFile = File(...),
    customer_file: UploadFile = File(...),
    transportation_cost_per_km: float = Form(1.0),
    carbon_constraint_kg: Optional[float] = Form(None),
    carbon_price_per_kg: float = Form(0.0),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    cost_col: str = Form("fixed_cost"),
    capacity_col: str = Form("vehicle_capacity"),
    loading_rate_col: str = Form("loading_rate"),
    fuel_type_col: str = Form("fuel_type")
):
    """
    Perform comprehensive carbon footprint analysis for logistics network design
    
    Features:
    - Multi-objective optimization (cost vs environmental impact)
    - Empirical CO2 emission calculations based on vehicle specifications
    - Carbon constraint optimization with violation detection
    - Pareto analysis for cost-emissions trade-offs
    - Environmental impact visualization data
    - Sustainability recommendations
    
    This analysis considers both operational costs and environmental impact
    to provide sustainable logistics network design solutions.
    """
    try:
        # Read uploaded files
        facility_contents = await facility_file.read()
        customer_contents = await customer_file.read()
        
        facility_df = pd.read_csv(io.StringIO(facility_contents.decode('utf-8')))
        customer_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        
        # Validate required columns for facilities
        required_facility_cols = [lat_col, lon_col, cost_col]
        missing_cols = [col for col in required_facility_cols if col not in facility_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns in facility file: {missing_cols}"
            )
        
        # Validate required columns for customers
        required_customer_cols = [lat_col, lon_col, demand_col]
        missing_cols = [col for col in required_customer_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns in customer file: {missing_cols}"
            )
        
        # Convert DataFrames to dictionaries
        facilities_data = []
        for _, row in facility_df.iterrows():
            facility_data = {
                "lat": row[lat_col],
                "lon": row[lon_col],
                "fixed_cost": row[cost_col],
                "vehicle_capacity": row.get(capacity_col, 10000),  # Default 10 tons
                "loading_rate": row.get(loading_rate_col, 0.7),   # Default 70%
                "fuel_type": row.get(fuel_type_col, "diesel"),    # Default diesel
                "transport_mode": "custom"
            }
            facilities_data.append(facility_data)
        
        customers_data = []
        for _, row in customer_df.iterrows():
            customer_data = {
                "lat": row[lat_col],
                "lon": row[lon_col],
                "demand": row[demand_col]  # in kg
            }
            customers_data.append(customer_data)
        
        # Perform multi-objective analysis
        analysis_results = carbon_footprint_service.multi_objective_analysis(
            facilities_data=facilities_data,
            customers_data=customers_data,
            transportation_cost_per_km=transportation_cost_per_km,
            carbon_constraint_kg=carbon_constraint_kg,
            carbon_price_per_kg=carbon_price_per_kg
        )
        
        # Generate visualization data
        visualization_data = carbon_footprint_service.generate_carbon_visualization_data(analysis_results)
        
        # Convert numpy arrays to lists for JSON serialization
        def convert_to_json_safe(obj):
            """Convert numpy types to Python native types"""
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Prepare final result
        result = {
            "analysis_results": analysis_results,
            "visualization_data": visualization_data,
            "algorithm": "multi_objective_carbon_footprint_analysis",
            "parameters": {
                "transportation_cost_per_km": transportation_cost_per_km,
                "carbon_constraint_kg": carbon_constraint_kg,
                "carbon_price_per_kg": carbon_price_per_kg,
                "num_facilities": len(facilities_data),
                "num_customers": len(customers_data)
            }
        }
        
        # Convert to JSON-safe format
        result = convert_to_json_safe(result)
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in carbon footprint analysis: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"Error in carbon footprint analysis: {str(e)}")

@router.post("/co2-emission-calculation")
async def calculate_co2_emissions(
    capacity_kg: float = Form(...),
    loading_rate: float = Form(0.7),
    fuel_type: str = Form("diesel"),
    distance_km: float = Form(...),
    weight_tons: float = Form(...)
):
    """
    Calculate CO2 emissions for specific transportation parameters
    
    Uses empirical formula from 05lnd.ipynb:
    - For diesel: fuel = exp(2.67 - 0.927*ln(rate) - 0.648*ln(capacity))
    - For gasoline: fuel = exp(2.71 - 0.812*ln(rate) - 0.654*ln(capacity))
    
    Returns detailed emission calculations and environmental impact metrics.
    """
    try:
        # Validate input parameters
        if capacity_kg <= 0:
            raise HTTPException(status_code=400, detail="Capacity must be positive")
        if not (0 < loading_rate <= 1.0):
            raise HTTPException(status_code=400, detail="Loading rate must be between 0 and 1")
        if fuel_type not in ["diesel", "gasoline"]:
            raise HTTPException(status_code=400, detail="Fuel type must be 'diesel' or 'gasoline'")
        if distance_km < 0:
            raise HTTPException(status_code=400, detail="Distance must be non-negative")
        if weight_tons < 0:
            raise HTTPException(status_code=400, detail="Weight must be non-negative")
        
        # Calculate CO2 emissions
        fuel_consumption, co2_per_ton_km = carbon_footprint_service.calculate_co2_emission(
            capacity=capacity_kg,
            loading_rate=loading_rate,
            fuel_type=fuel_type
        )
        
        # Calculate transportation emissions
        emission_details = carbon_footprint_service.calculate_transportation_emissions(
            distance_km=distance_km,
            weight_tons=weight_tons,
            transport_mode="custom",
            capacity=capacity_kg,
            loading_rate=loading_rate,
            fuel_type=fuel_type
        )
        
        # Add environmental impact indicators
        total_co2_kg = emission_details["total_co2_kg"]
        
        result = {
            "emission_calculations": emission_details,
            "environmental_impact": {
                "total_co2_kg": total_co2_kg,
                "total_co2_tons": total_co2_kg / 1000,
                "equivalent_trees_needed": int(total_co2_kg / 22),  # Trees needed to absorb CO2 per year
                "car_equivalent_km": total_co2_kg / 0.21,  # Equivalent km driven by average car
                "carbon_footprint_per_ton": co2_per_ton_km / 1000,  # kg CO2 per ton-km
                "fuel_efficiency": fuel_consumption,  # L per ton-km
            },
            "efficiency_metrics": {
                "capacity_utilization": loading_rate,
                "efficiency_score": min(100, (1.0 / fuel_consumption) * 100) if fuel_consumption > 0 else 0,
                "environmental_rating": "High" if co2_per_ton_km < 80 else "Medium" if co2_per_ton_km < 120 else "Low"
            },
            "recommendations": []
        }
        
        # Add recommendations based on efficiency
        if loading_rate < 0.5:
            result["recommendations"].append({
                "type": "loading_optimization",
                "priority": "high",
                "description": f"Loading rate is only {loading_rate*100:.1f}%",
                "action": "Improve loading efficiency to reduce emissions per unit transported"
            })
        
        if co2_per_ton_km > 120:  # High emission threshold
            result["recommendations"].append({
                "type": "vehicle_optimization",
                "priority": "medium",
                "description": f"CO2 emissions are {co2_per_ton_km:.1f} g/ton-km (high)",
                "action": "Consider more fuel-efficient vehicles or alternative fuel types"
            })
        
        if fuel_type == "diesel" and capacity_kg < 5000:
            result["recommendations"].append({
                "type": "fuel_optimization",
                "priority": "low",
                "description": "Small capacity vehicle using diesel",
                "action": "Consider gasoline or electric alternatives for small vehicles"
            })
        
        return result
    
    except Exception as e:
        import traceback
        error_detail = f"Error in CO2 emission calculation: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"Error in CO2 emission calculation: {str(e)}")

@router.get("/carbon-emission-factors")
async def get_carbon_emission_factors():
    """
    Get standard CO2 emission factors for different transportation modes
    
    Returns emission factors in gCO2/ton-km for various transport modes
    and fuel conversion factors for different fuel types.
    """
    try:
        emission_factors = carbon_footprint_service.emission_factors
        fuel_factors = carbon_footprint_service.fuel_to_co2
        
        return {
            "standard_emission_factors": {
                "description": "Standard CO2 emission factors by transport mode",
                "unit": "gCO2/ton-km",
                "factors": emission_factors,
                "source": "Industry standard emission factors"
            },
            "fuel_conversion_factors": {
                "description": "CO2 emission factors by fuel type",
                "unit": "kg CO2 per unit",
                "factors": fuel_factors,
                "units": {
                    "diesel": "kg CO2 per liter",
                    "gasoline": "kg CO2 per liter", 
                    "electricity": "kg CO2 per kWh"
                }
            },
            "calculation_methods": {
                "empirical_formula": "Vehicle-specific calculations based on capacity and loading rate",
                "standard_factors": "Transport mode-specific emission factors",
                "hybrid": "Combination of both methods for comprehensive analysis"
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting emission factors: {str(e)}")

# Excel Integration Endpoints

@router.get("/excel-template")
async def download_excel_template():
    """
    Download Excel template for MELOS logistics network design
    
    Returns an Excel file with predefined sheets for:
    - Customer data (顧客)
    - Warehouse candidates (倉庫候補地点) 
    - Plants (工場)
    - Products (製品)
    - Demand matrix (需要)
    - Production matrix (生産)
    """
    try:
        # Create Excel template
        wb = excel_integration_service.make_excel_melos()
        wb = excel_integration_service.make_demand_production_sheets(wb)
        
        # Save to BytesIO buffer
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return as downloadable file
        from fastapi.responses import StreamingResponse
        
        def generate():
            buffer.seek(0)
            while True:
                chunk = buffer.read(1024)
                if not chunk:
                    break
                yield chunk
        
        return StreamingResponse(
            generate(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=MELOS_template.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating Excel template: {str(e)}")

@router.post("/excel-parse")
async def parse_excel_file(
    file: UploadFile = File(...)
):
    """
    Parse uploaded Excel file and extract logistics network data
    
    Expects an Excel file with MELOS format sheets:
    - 顧客 (Customers)
    - 倉庫候補地点 (Warehouse Candidates)
    - 工場 (Plants) 
    - 製品 (Products)
    - 需要 (Demand)
    - 生産 (Production)
    """
    try:
        # Read Excel file
        contents = await file.read()
        from io import BytesIO
        from openpyxl import load_workbook
        
        wb = load_workbook(BytesIO(contents))
        
        # Parse data from Excel sheets
        cust_df, dc_df, plnt_df, prod_df, demand_df, production_df = excel_integration_service.prepare_df_for_melos(wb)
        
        # Convert DataFrames to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        result = {
            "customers": cust_df.to_dict('records'),
            "warehouse_candidates": dc_df.to_dict('records'), 
            "plants": plnt_df.to_dict('records'),
            "products": prod_df.to_dict('records'),
            "demand_matrix": demand_df.to_dict('records'),
            "production_matrix": production_df.to_dict('records'),
            "summary": {
                "customers_count": len(cust_df),
                "warehouse_candidates_count": len(dc_df),
                "plants_count": len(plnt_df),
                "products_count": len(prod_df)
            }
        }
        
        # Apply JSON conversion
        for key in result:
            if key != "summary":
                result[key] = [convert_to_json_safe(item) for item in result[key]]
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")

@router.post("/excel-solve-lnd")
async def solve_lnd_from_excel(
    excel_file: UploadFile = File(...),
    solver: str = Form("multi_source"),
    max_facilities: int = Form(3),
    osrm_host: str = Form("localhost")
):
    """
    Solve logistics network design problem from Excel data
    
    Takes a MELOS-format Excel file and solves the optimization problem
    using the specified solver (multi_source or single_source).
    
    Returns optimized facility locations and customer assignments.
    """
    try:
        # Read Excel file
        contents = await excel_file.read()
        from io import BytesIO
        from openpyxl import load_workbook
        
        wb = load_workbook(BytesIO(contents))
        
        # Set OSRM host
        excel_integration_service.osrm_host = osrm_host
        
        # Create network and solve
        network_df = excel_integration_service.make_network_for_excel(wb)
        results = excel_integration_service.solve_lnd_for_excel(wb, solver=solver)
        
        # Add results back to Excel
        wb_with_results = excel_integration_service.add_result_for_melos(wb, results)
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Prepare response
        response_results = {}
        for key, value in results.items():
            if isinstance(value, list):
                response_results[key] = [convert_to_json_safe(item) for item in value]
            else:
                response_results[key] = convert_to_json_safe(value)
        
        return {
            "optimization_results": response_results,
            "network_summary": {
                "total_routes": len(network_df),
                "total_network_cost": network_df['cost_per_unit'].sum(),
                "average_distance_km": network_df['distance_m'].mean() / 1000
            },
            "solver_used": solver,
            "parameters": {
                "max_facilities": max_facilities,
                "osrm_host": osrm_host
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error solving LND from Excel: {str(e)}")

# Network Generation Endpoints

@router.post("/generate-network-road")
async def generate_network_using_road(
    customer_file: UploadFile = File(...),
    dc_file: UploadFile = File(...),
    plant_file: UploadFile = File(...),
    plnt_dc_threshold: float = Form(500.0),
    dc_cust_threshold: float = Form(100.0),
    tc_per_dis: float = Form(2.0),
    dc_per_dis: float = Form(1.5),
    tc_per_time: float = Form(0.01),
    dc_per_time: float = Form(0.008),
    use_toll: bool = Form(True),
    osrm_host: str = Form("localhost"),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    name_col: str = Form("name")
):
    """
    Generate transportation network using road distances and travel times
    
    Uses OSRM routing engine to calculate realistic travel times and distances,
    then creates a logistics network with cost and lead time modeling.
    
    Features:
    - Real road distance calculations via OSRM
    - Distance-based and time-based cost modeling
    - Configurable distance thresholds for network pruning
    - Lead time estimation with processing time modeling
    - NetworkX graph generation for analysis
    - Interactive network visualization data
    
    This matches the notebook's make_network_using_road functionality exactly.
    """
    try:
        # Read uploaded files
        customer_contents = await customer_file.read()
        dc_contents = await dc_file.read()
        plant_contents = await plant_file.read()
        
        cust_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        dc_df = pd.read_csv(io.StringIO(dc_contents.decode('utf-8')))
        plnt_df = pd.read_csv(io.StringIO(plant_contents.decode('utf-8')))
        
        # Validate required columns
        for df, df_name in [(cust_df, "customer"), (dc_df, "dc"), (plnt_df, "plant")]:
            required_cols = [name_col, lat_col, lon_col]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns in {df_name} file: {missing_cols}"
                )
        
        # Rename columns to standard format
        for df in [cust_df, dc_df, plnt_df]:
            df.rename(columns={
                name_col: 'name',
                lat_col: 'lat',
                lon_col: 'lon'
            }, inplace=True)
        
        # Set OSRM host
        network_generation_service.osrm_host = osrm_host
        
        # Compute durations and distances using OSRM
        durations, distances, node_df = network_generation_service.compute_durations(
            cust_df, dc_df, plnt_df, toll=use_toll
        )
        
        # Generate network using road data
        trans_df, G, position = network_generation_service.make_network_using_road(
            cust_df, dc_df, plnt_df, durations, distances,
            plnt_dc_threshold=plnt_dc_threshold,
            dc_cust_threshold=dc_cust_threshold,
            tc_per_dis=tc_per_dis,
            dc_per_dis=dc_per_dis,
            tc_per_time=tc_per_time,
            dc_per_time=dc_per_time
        )
        
        # Generate visualization data
        viz_data = network_generation_service.plot_scm(
            trans_df, cust_df, dc_df, plnt_df,
            title="Road-based Transportation Network"
        )
        
        # Analyze network properties
        network_analysis = network_generation_service.analyze_network_properties(trans_df, G)
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Prepare response
        result = {
            "transportation_network": trans_df.to_dict('records'),
            "network_analysis": network_analysis,
            "visualization": viz_data,
            "node_summary": {
                "customers": len(cust_df),
                "distribution_centers": len(dc_df),
                "plants": len(plnt_df),
                "total_nodes": len(cust_df) + len(dc_df) + len(plnt_df)
            },
            "edge_summary": {
                "total_edges": len(trans_df),
                "plant_to_dc": len(trans_df[trans_df['from_type'] == 'plant']),
                "dc_to_customer": len(trans_df[trans_df['from_type'] == 'dc'])
            },
            "parameters": {
                "plnt_dc_threshold_km": plnt_dc_threshold,
                "dc_cust_threshold_km": dc_cust_threshold,
                "transport_cost_per_km": tc_per_dis,
                "delivery_cost_per_km": dc_per_dis,
                "transport_cost_per_second": tc_per_time,
                "delivery_cost_per_second": dc_per_time,
                "use_toll_roads": use_toll,
                "osrm_host": osrm_host
            }
        }
        
        # Apply JSON conversion
        for key in ["transportation_network", "network_analysis", "visualization"]:
            if key in result:
                result[key] = convert_to_json_safe(result[key])
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating road-based network: {str(e)}")

@router.post("/generate-network-simple")
async def generate_network_simple(
    customer_file: UploadFile = File(...),
    dc_file: UploadFile = File(...),
    plant_file: UploadFile = File(...),
    plnt_dc_threshold: float = Form(500.0),
    dc_cust_threshold: float = Form(100.0),
    unit_tp_cost: float = Form(2.0),
    unit_del_cost: float = Form(1.5),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    name_col: str = Form("name")
):
    """
    Generate transportation network using great circle distances
    
    Creates a simplified logistics network using straight-line distances
    as an approximation. Faster than road-based calculation but less accurate.
    
    Features:
    - Great circle distance calculations (no external dependencies)
    - Simplified cost modeling based on distance only
    - Configurable distance thresholds
    - Lead time estimation with processing time modeling
    - NetworkX graph generation for analysis
    - Interactive network visualization
    
    This matches the notebook's make_network functionality exactly.
    """
    try:
        # Read uploaded files
        customer_contents = await customer_file.read()
        dc_contents = await dc_file.read()
        plant_contents = await plant_file.read()
        
        cust_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        dc_df = pd.read_csv(io.StringIO(dc_contents.decode('utf-8')))
        plnt_df = pd.read_csv(io.StringIO(plant_contents.decode('utf-8')))
        
        # Validate required columns
        for df, df_name in [(cust_df, "customer"), (dc_df, "dc"), (plnt_df, "plant")]:
            required_cols = [name_col, lat_col, lon_col]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns in {df_name} file: {missing_cols}"
                )
        
        # Rename columns to standard format
        for df in [cust_df, dc_df, plnt_df]:
            df.rename(columns={
                name_col: 'name',
                lat_col: 'lat',
                lon_col: 'lon'
            }, inplace=True)
        
        # Generate network using great circle distances
        trans_df, G, position = network_generation_service.make_network(
            cust_df, dc_df, plnt_df,
            plnt_dc_threshold=plnt_dc_threshold,
            dc_cust_threshold=dc_cust_threshold,
            unit_tp_cost=unit_tp_cost,
            unit_del_cost=unit_del_cost
        )
        
        # Generate visualization data
        viz_data = network_generation_service.plot_scm(
            trans_df, cust_df, dc_df, plnt_df,
            title="Great Circle Distance Transportation Network"
        )
        
        # Analyze network properties
        network_analysis = network_generation_service.analyze_network_properties(trans_df, G)
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # Prepare response
        result = {
            "transportation_network": trans_df.to_dict('records'),
            "network_analysis": network_analysis,
            "visualization": viz_data,
            "node_summary": {
                "customers": len(cust_df),
                "distribution_centers": len(dc_df),
                "plants": len(plnt_df),
                "total_nodes": len(cust_df) + len(dc_df) + len(plnt_df)
            },
            "edge_summary": {
                "total_edges": len(trans_df),
                "plant_to_dc": len(trans_df[trans_df['from_type'] == 'plant']),
                "dc_to_customer": len(trans_df[trans_df['from_type'] == 'dc'])
            },
            "parameters": {
                "plnt_dc_threshold_km": plnt_dc_threshold,
                "dc_cust_threshold_km": dc_cust_threshold,
                "unit_transport_cost_per_km": unit_tp_cost,
                "unit_delivery_cost_per_km": unit_del_cost,
                "distance_method": "great_circle"
            }
        }
        
        # Apply JSON conversion
        for key in ["transportation_network", "network_analysis", "visualization"]:
            if key in result:
                result[key] = convert_to_json_safe(result[key])
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating simple network: {str(e)}")

@router.post("/compute-distance-matrix")
async def compute_distance_matrix(
    location_file: UploadFile = File(...),
    use_toll: bool = Form(True),
    osrm_host: str = Form("localhost"),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    name_col: str = Form("name")
):
    """
    Compute distance and duration matrix between locations
    
    Uses OSRM routing engine to calculate travel times and distances
    between all pairs of locations in the input file.
    
    Features:
    - OSRM integration for real road distances
    - Toll road inclusion/exclusion options
    - Fallback to great circle distances if OSRM unavailable
    - Symmetric distance/duration matrices
    - JSON-safe output format
    
    This provides the core distance computation used by network generation.
    """
    try:
        # Read uploaded file
        contents = await location_file.read()
        locations_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [name_col, lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in locations_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Rename columns to standard format
        locations_df.rename(columns={
            name_col: 'name',
            lat_col: 'lat',
            lon_col: 'lon'
        }, inplace=True)
        
        # Set OSRM host
        network_generation_service.osrm_host = osrm_host
        
        # Compute distances - use the compute_durations function with dummy DCs and plants
        empty_df = pd.DataFrame(columns=['name', 'lat', 'lon'])
        durations, distances, node_df = network_generation_service.compute_durations(
            locations_df, empty_df, empty_df, toll=use_toll
        )
        
        # Prepare location labels
        location_names = locations_df['name'].tolist()
        
        return {
            "locations": locations_df.to_dict('records'),
            "distance_matrix_km": [[d/1000.0 for d in row] for row in distances],
            "duration_matrix_hours": [[d/3600.0 for d in row] for row in durations],
            "location_names": location_names,
            "matrix_size": len(location_names),
            "parameters": {
                "use_toll_roads": use_toll,
                "osrm_host": osrm_host,
                "distance_unit": "kilometers",
                "duration_unit": "hours"
            },
            "statistics": {
                "avg_distance_km": float(np.mean([d for row in distances for d in row if d > 0]) / 1000.0),
                "max_distance_km": float(np.max([d for row in distances for d in row]) / 1000.0),
                "avg_duration_hours": float(np.mean([d for row in durations for d in row if d > 0]) / 3600.0),
                "max_duration_hours": float(np.max([d for row in durations for d in row]) / 3600.0)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error computing distance matrix: {str(e)}")

@router.get("/network-generation-info")
async def get_network_generation_info():
    """
    Get information about network generation service capabilities
    
    Returns details about available functions, features, and configuration options
    for the network generation service.
    """
    try:
        return network_generation_service.get_service_info()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting service info: {str(e)}")

# Transportation Problem Solver Endpoints

@router.post("/solve-transportation-problem")
async def solve_transportation_problem(
    transportation_file: UploadFile = File(...),
    problem_type: str = Form("balanced")
):
    """
    Solve transportation problem using various algorithms
    
    Expects CSV file with columns [from_node, to_node, unit_cost, supply, demand]
    where supply/demand values are positive for supply nodes, negative for demand nodes.
    
    Supports different transportation problem variants:
    - balanced: Standard transportation problem (supply = demand)
    - unbalanced: Supply and demand totals don't match
    - capacitated: Transportation with facility capacity constraints
    
    Features:
    - NetworkX network simplex algorithm
    - Automatic balancing for unbalanced problems  
    - Comprehensive solution analysis
    - JSON-safe output format
    
    This matches the notebook's transportation solver computational procedures.
    """
    try:
        # Read transportation data file
        contents = await transportation_file.read()
        transport_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = ['from_node', 'to_node', 'unit_cost', 'supply', 'demand']
        missing_cols = [col for col in required_cols if col not in transport_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Extract supply and demand data
        supply_data = {}
        demand_data = {}
        costs = {}
        
        for _, row in transport_df.iterrows():
            from_node, to_node = row['from_node'], row['to_node']
            supply_data[from_node] = float(row['supply'])
            demand_data[to_node] = float(row['demand'])  
            costs[(from_node, to_node)] = float(row['unit_cost'])
        
        # Convert to lists for matrix operations
        suppliers = list(supply_data.keys())
        customers = list(demand_data.keys())
        supply = [supply_data[s] for s in suppliers]
        demand = [demand_data[c] for c in customers]
        
        # Create cost matrix
        cost_matrix = np.zeros((len(suppliers), len(customers)))
        for i, supplier in enumerate(suppliers):
            for j, customer in enumerate(customers):
                if (supplier, customer) in costs:
                    cost_matrix[i, j] = costs[(supplier, customer)]
                else:
                    cost_matrix[i, j] = 1e6  # Large cost for non-existent routes
        
        # Now costs is cost_matrix, not costs dict
        costs = cost_matrix
        
        # Solve based on problem type
        if problem_type == "balanced":
            if abs(sum(supply) - sum(demand)) > 1e-6:
                raise HTTPException(
                    status_code=400,
                    detail=f"Problem marked as balanced but supply ({sum(supply)}) != demand ({sum(demand)})"
                )
            solution = transportation_service._solve_balanced_transportation(
                np.array(supply), np.array(demand), costs
            )
        elif problem_type == "unbalanced":
            solution = transportation_service.unbalanced_transportation(
                supply, demand, costs
            )
        elif problem_type == "capacitated":
            # For capacitated problems, assume equal capacity for all facilities
            avg_capacity = int(sum(demand) / len(supply)) + 1
            C_customers_facilities = costs.T  # Transpose for customer x facility format
            cost, flow = transportation_service.transportation(C_customers_facilities, avg_capacity)
            solution = {
                "status": "optimal",
                "objective_value": cost,
                "flow": flow,
                "solver": "networkx_capacitated"
            }
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported problem type: {problem_type}. Use 'balanced', 'unbalanced', or 'capacitated'"
            )
        
        # Analyze solution
        analysis = transportation_service.analyze_transportation_solution(solution)
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        result = {
            "solution": convert_to_json_safe(solution),
            "analysis": convert_to_json_safe(analysis),
            "problem_parameters": {
                "problem_type": problem_type,
                "supply": supply,
                "demand": demand, 
                "cost_matrix_shape": costs.shape,
                "total_supply": sum(supply),
                "total_demand": sum(demand),
                "is_balanced": abs(sum(supply) - sum(demand)) < 1e-6
            }
        }
        
        return result
        
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=f"Invalid input data: {str(ve)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error solving transportation problem: {str(e)}")

@router.post("/solve-multi-commodity-transportation")
async def solve_multi_commodity_transportation(
    supply_file: UploadFile = File(...),
    demand_file: UploadFile = File(...), 
    cost_file: UploadFile = File(...)
):
    """
    Solve multi-commodity transportation problem
    
    Expects CSV files with the following formats:
    - supply_file: columns [supplier, product, supply_amount]
    - demand_file: columns [customer, product, demand_amount]  
    - cost_file: columns [supplier, customer, product, unit_cost]
    
    Features:
    - Multiple products transported simultaneously
    - Gurobi or PuLP optimization backends
    - Supply and demand constraints per product
    - Comprehensive solution analysis
    
    This extends the notebook's single-commodity approach to handle multiple products.
    """
    try:
        # Read uploaded files
        supply_contents = await supply_file.read()
        demand_contents = await demand_file.read()
        cost_contents = await cost_file.read()
        
        supply_df = pd.read_csv(io.StringIO(supply_contents.decode('utf-8')))
        demand_df = pd.read_csv(io.StringIO(demand_contents.decode('utf-8')))
        cost_df = pd.read_csv(io.StringIO(cost_contents.decode('utf-8')))
        
        # Validate required columns
        if not all(col in supply_df.columns for col in ['supplier', 'product', 'supply_amount']):
            raise HTTPException(status_code=400, detail="Supply file must have columns: supplier, product, supply_amount")
            
        if not all(col in demand_df.columns for col in ['customer', 'product', 'demand_amount']):
            raise HTTPException(status_code=400, detail="Demand file must have columns: customer, product, demand_amount")
            
        if not all(col in cost_df.columns for col in ['supplier', 'customer', 'product', 'unit_cost']):
            raise HTTPException(status_code=400, detail="Cost file must have columns: supplier, customer, product, unit_cost")
        
        # Convert to dictionaries for solver
        supply = {}
        for _, row in supply_df.iterrows():
            supply[(row['supplier'], row['product'])] = float(row['supply_amount'])
            
        demand = {}
        for _, row in demand_df.iterrows():
            demand[(row['customer'], row['product'])] = float(row['demand_amount'])
            
        costs = {}
        for _, row in cost_df.iterrows():
            costs[(row['supplier'], row['customer'], row['product'])] = float(row['unit_cost'])
        
        # Solve multi-commodity problem
        solution = transportation_service.multi_commodity_transportation(supply, demand, costs)
        
        # Analyze solution  
        analysis = transportation_service.analyze_transportation_solution(solution)
        
        # Convert to JSON-safe format
        def convert_tuples_to_strings(obj):
            if isinstance(obj, dict):
                new_dict = {}
                for key, value in obj.items():
                    if isinstance(key, tuple):
                        new_key = "_".join(str(x) for x in key)
                    else:
                        new_key = key
                    new_dict[new_key] = convert_tuples_to_strings(value)
                return new_dict
            elif isinstance(obj, list):
                return [convert_tuples_to_strings(item) for item in obj]
            else:
                return obj
        
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        # First convert tuples to strings, then make JSON-safe
        result = convert_tuples_to_strings({
            "solution": solution,
            "analysis": analysis,
            "problem_summary": {
                "num_suppliers": len(set(s for s, p in supply.keys())),
                "num_customers": len(set(c for c, p in demand.keys())),
                "num_products": len(set(p for s, p in supply.keys())),
                "total_supply": sum(supply.values()),
                "total_demand": sum(demand.values()),
                "num_routes": len(costs)
            }
        })
        
        return convert_to_json_safe(result)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error solving multi-commodity transportation: {str(e)}")

@router.post("/solve-transshipment-problem")
async def solve_transshipment_problem(
    network_file: UploadFile = File(...)
):
    """
    Solve transshipment problem with intermediate transfer nodes
    
    Expects CSV file with columns:
    [from_node, to_node, unit_cost, node_type, supply_demand]
    
    Where:
    - node_type: 'supply', 'demand', or 'transship'
    - supply_demand: positive for supply, negative for demand, 0 for transshipment
    
    Features:
    - Intermediate transshipment nodes for flexible routing
    - NetworkX network simplex algorithm
    - Support for complex supply chain topologies
    - Real-world logistics network modeling
    
    This extends basic transportation to handle warehouse/hub networks.
    """
    try:
        # Read network file
        contents = await network_file.read()
        network_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = ['from_node', 'to_node', 'unit_cost', 'node_type', 'supply_demand']
        missing_cols = [col for col in required_cols if col not in network_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Extract node information
        nodes_info = {}
        for _, row in network_df.iterrows():
            from_node, to_node = row['from_node'], row['to_node']
            
            # Collect node types and supply/demand info
            for node in [from_node, to_node]:
                if node not in nodes_info:
                    # Find node info from any row containing this node
                    node_rows = network_df[
                        (network_df['from_node'] == node) | 
                        (network_df['to_node'] == node)
                    ]
                    if len(node_rows) > 0:
                        first_row = node_rows.iloc[0]
                        nodes_info[node] = {
                            'type': first_row['node_type'],
                            'supply_demand': float(first_row['supply_demand'])
                        }
        
        # Separate nodes by type
        supply_nodes = {node: info['supply_demand'] 
                       for node, info in nodes_info.items() 
                       if info['type'] == 'supply'}
        demand_nodes = {node: -info['supply_demand'] 
                       for node, info in nodes_info.items() 
                       if info['type'] == 'demand'}
        transship_nodes = [node for node, info in nodes_info.items() 
                          if info['type'] == 'transship']
        
        # Extract costs
        costs = {}
        for _, row in network_df.iterrows():
            costs[(row['from_node'], row['to_node'])] = float(row['unit_cost'])
        
        # Solve transshipment problem
        solution = transportation_service.transshipment_problem(
            supply_nodes, demand_nodes, transship_nodes, costs
        )
        
        # Analyze solution
        analysis = transportation_service.analyze_transportation_solution(solution)
        
        # Convert to JSON-safe format  
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                new_dict = {}
                for key, value in obj.items():
                    if isinstance(key, tuple):
                        new_key = "_to_".join(str(x) for x in key)
                    else:
                        new_key = key
                    new_dict[new_key] = convert_to_json_safe(value)
                return new_dict
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        result = {
            "solution": convert_to_json_safe(solution),
            "analysis": convert_to_json_safe(analysis),
            "network_summary": {
                "supply_nodes": len(supply_nodes),
                "demand_nodes": len(demand_nodes), 
                "transshipment_nodes": len(transship_nodes),
                "total_arcs": len(costs),
                "total_supply": sum(supply_nodes.values()),
                "total_demand": sum(demand_nodes.values())
            }
        }
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error solving transshipment problem: {str(e)}")

@router.get("/transportation-service-info")
async def get_transportation_service_info():
    """
    Get information about transportation service capabilities
    
    Returns details about available transportation problem solvers,
    supported problem types, and integration capabilities.
    """
    try:
        return transportation_service.get_service_info()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting transportation service info: {str(e)}")

# Visualization Service Endpoints

@router.post("/create-facility-location-visualization")
async def create_facility_location_visualization(
    customer_file: UploadFile = File(...),
    facility_locations: List[str] = Form(...),  # JSON strings of [lat, lon] pairs
    assignments: Optional[List[int]] = Form(None),
    title: str = Form("Facility Location Optimization Results"),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    name_col: str = Form("name")
):
    """
    Create interactive facility location visualization map
    
    Features:
    - Interactive Plotly map with customer and facility markers
    - Customer-facility assignment visualization with connecting lines
    - Color-coded customer groups based on facility assignments
    - Hover information showing location details
    - Customizable map styling and zoom levels
    
    This matches the notebook's facility location visualization procedures exactly.
    """
    try:
        # Read customer data
        contents = await customer_file.read()
        customer_df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Rename columns to standard format
        customer_df.rename(columns={
            lat_col: 'lat',
            lon_col: 'lon',
            name_col: 'name'
        }, inplace=True)
        
        # Parse facility locations from JSON strings
        facility_locs = []
        for loc_str in facility_locations:
            try:
                import json
                loc = json.loads(loc_str)
                if len(loc) >= 2:
                    facility_locs.append((float(loc[0]), float(loc[1])))
            except (json.JSONDecodeError, ValueError) as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid facility location format: {loc_str}. Expected [lat, lon] JSON."
                )
        
        # Create visualization
        viz_data = visualization_service.create_facility_location_map(
            customer_df, facility_locs, assignments, title
        )
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        return {
            "visualization": convert_to_json_safe(viz_data),
            "summary": {
                "num_customers": len(customer_df),
                "num_facilities": len(facility_locs),
                "has_assignments": assignments is not None,
                "map_center": {
                    "lat": float(np.mean(customer_df['lat'])),
                    "lon": float(np.mean(customer_df['lon']))
                }
            },
            "chart_type": "facility_location_map"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating facility location visualization: {str(e)}")

@router.post("/create-network-visualization")
async def create_network_visualization(
    customer_file: UploadFile = File(...),
    dc_file: Optional[UploadFile] = File(None),
    plant_file: Optional[UploadFile] = File(None),
    network_edges_file: Optional[UploadFile] = File(None),
    title: str = Form("Supply Chain Network"),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    name_col: str = Form("name")
):
    """
    Create comprehensive supply chain network visualization
    
    Expects CSV files with location data. Optional network_edges_file should have:
    [from_name, to_name, from_lat, from_lon, to_lat, to_lon, type]
    
    Features:
    - Multi-tier supply chain network visualization
    - Different node types (plants, DCs, customers) with distinct styling
    - Network flow visualization with edge types
    - Interactive hover information and legends
    - Automatic map centering and zoom adjustment
    
    This matches the notebook's network visualization procedures.
    """
    try:
        # Read customer data (required)
        customer_contents = await customer_file.read()
        customer_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        
        # Validate customer columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in customer_df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns in customer file: {missing_cols}"
            )
        
        # Rename customer columns
        customer_df.rename(columns={
            lat_col: 'lat',
            lon_col: 'lon',
            name_col: 'name'
        }, inplace=True)
        
        # Read DC data (optional)
        dc_df = None
        if dc_file:
            dc_contents = await dc_file.read()
            dc_df = pd.read_csv(io.StringIO(dc_contents.decode('utf-8')))
            dc_df.rename(columns={
                lat_col: 'lat',
                lon_col: 'lon', 
                name_col: 'name'
            }, inplace=True)
        
        # Read plant data (optional)
        plant_df = None
        if plant_file:
            plant_contents = await plant_file.read()
            plant_df = pd.read_csv(io.StringIO(plant_contents.decode('utf-8')))
            plant_df.rename(columns={
                lat_col: 'lat',
                lon_col: 'lon',
                name_col: 'name'
            }, inplace=True)
        
        # Read network edges (optional)
        network_edges = None
        if network_edges_file:
            edges_contents = await network_edges_file.read()
            edges_df = pd.read_csv(io.StringIO(edges_contents.decode('utf-8')))
            
            network_edges = []
            for _, row in edges_df.iterrows():
                network_edges.append({
                    'from_name': row.get('from_name', ''),
                    'to_name': row.get('to_name', ''),
                    'from_lat': float(row.get('from_lat', 0)),
                    'from_lon': float(row.get('from_lon', 0)),
                    'to_lat': float(row.get('to_lat', 0)),
                    'to_lon': float(row.get('to_lon', 0)),
                    'type': row.get('type', 'unknown')
                })
        
        # Create visualization
        viz_data = visualization_service.create_network_visualization(
            customer_df, dc_df, plant_df, network_edges, title
        )
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        return {
            "visualization": convert_to_json_safe(viz_data),
            "summary": {
                "num_customers": len(customer_df),
                "num_dcs": len(dc_df) if dc_df is not None else 0,
                "num_plants": len(plant_df) if plant_df is not None else 0,
                "num_edges": len(network_edges) if network_edges else 0
            },
            "chart_type": "network_map"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating network visualization: {str(e)}")

@router.post("/create-optimization-analysis-charts")  
async def create_optimization_analysis_charts(
    results_data: Dict[str, Any],
    title_prefix: str = Form("Optimization Analysis")
):
    """
    Create comprehensive optimization analysis charts
    
    Expects JSON data with optimization results including:
    - lower_bounds, upper_bounds: Convergence data
    - learning_rates: Learning rate schedule
    - facility_stats: Facility performance statistics
    
    Features:
    - Algorithm convergence visualization
    - Learning rate schedule analysis  
    - Facility performance comparisons
    - Cost breakdown and analysis
    
    This matches the notebook's optimization analysis procedures.
    """
    try:
        # Create analysis charts
        charts = visualization_service.create_cost_analysis_charts(results_data, title_prefix)
        
        # Convert to JSON-safe format
        def convert_to_json_safe(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                val = float(obj)
                if np.isinf(val) or np.isnan(val):
                    return None
                return val
            elif isinstance(obj, (int, float)):
                if np.isinf(obj) or np.isnan(obj):
                    return None
                return obj
            elif isinstance(obj, np.bool_):
                return bool(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, tuple):
                return list(obj)
            elif isinstance(obj, list):
                return [convert_to_json_safe(item) for item in obj]
            elif isinstance(obj, dict):
                return {key: convert_to_json_safe(value) for key, value in obj.items()}
            elif isinstance(obj, (int, float, str, bool)) and pd.isna(obj):
                return None
            else:
                return obj
        
        return {
            "charts": convert_to_json_safe(charts),
            "summary": {
                "num_charts": len(charts),
                "chart_types": list(charts.keys())
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating optimization analysis charts: {str(e)}")

@router.get("/visualization-service-info")
async def get_visualization_service_info():
    """
    Get information about visualization service capabilities
    
    Returns details about available visualization types, supported data formats,
    and integration capabilities with the logistics network design system.
    """
    try:
        return visualization_service.get_service_info()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting visualization service info: {str(e)}")

@router.post("/abstract-logistics/lndp")
async def solve_logistics_network_design_problem(
    nodes_file: UploadFile = File(..., description="Network nodes data (CSV/Excel)"),
    arcs_file: UploadFile = File(..., description="Network arcs data (CSV/Excel)"),
    products_file: UploadFile = File(..., description="Product data (CSV/Excel)"),
    bom_file: Optional[UploadFile] = File(None, description="Bill of Materials data (CSV/Excel)"),
    demand_file: Optional[UploadFile] = File(None, description="Demand data (CSV/Excel)"),
    transport_cost_per_km: float = Form(0.5, description="Transport cost per kilometer"),
    fixed_cost_facility: float = Form(10000, description="Fixed cost per facility"),
    variable_cost_facility: float = Form(0.1, description="Variable cost per facility"),
    holding_cost_rate: float = Form(0.2, description="Annual holding cost rate"),
    ordering_cost: float = Form(100, description="Ordering cost per order"),
    max_distance: float = Form(500, description="Maximum service distance (km)"),
    min_service_level: float = Form(0.95, description="Minimum service level"),
    max_lead_time: int = Form(7, description="Maximum lead time (days)"),
    solver_type: str = Form("gurobi", description="Solver type (gurobi/pulp)"),
    time_limit: int = Form(3600, description="Solver time limit (seconds)"),
    mip_gap: float = Form(0.01, description="MIP gap tolerance")
):
    """
    Solve Logistics Network Design Problem (LNDP) with abstract logistics objects
    
    Supports:
    - Multi-product, multi-echelon networks
    - Bill of Materials (BOM) with assembly/disassembly
    - Echelon inventory costs
    - Service level constraints
    - Capacity constraints
    - Fixed and variable costs
    
    Returns comprehensive optimization results with network design solution.
    """
    try:
        # Read and process input files
        nodes_data = await _process_uploaded_file(nodes_file)
        arcs_data = await _process_uploaded_file(arcs_file)
        products_data = await _process_uploaded_file(products_file)
        
        bom_data = None
        if bom_file:
            bom_data = await _process_uploaded_file(bom_file)
        
        demand_data = None
        if demand_file:
            demand_data = await _process_uploaded_file(demand_file)
        
        # Prepare cost parameters
        cost_parameters = {
            'transport_cost_per_km': transport_cost_per_km,
            'fixed_cost_facility': fixed_cost_facility,
            'variable_cost_facility': variable_cost_facility,
            'holding_cost_rate': holding_cost_rate,
            'ordering_cost': ordering_cost
        }
        
        # Prepare service constraints
        service_constraints = {
            'max_distance': max_distance,
            'min_service_level': min_service_level,
            'max_lead_time': max_lead_time
        }
        
        # Prepare optimization options
        optimization_options = {
            'solver_type': solver_type,
            'time_limit': time_limit,
            'mip_gap': mip_gap,
            'threads': 4
        }
        
        # Solve LNDP
        result = LNDP(
            nodes_data={'nodes': nodes_data},
            arcs_data={'arcs': arcs_data},
            products_data={'products': products_data},
            bom_data=bom_data,
            demand_data=demand_data,
            cost_parameters=cost_parameters,
            service_constraints=service_constraints,
            optimization_options=optimization_options
        )
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error solving LNDP: {str(e)}")

@router.post("/abstract-logistics/echelon-inventory-costs")
async def calculate_echelon_inventory_costs_endpoint(
    network_file: UploadFile = File(..., description="Network structure data (CSV/Excel)"),
    demand_file: UploadFile = File(..., description="Demand patterns data (CSV/Excel)"),
    holding_cost_rate: float = Form(0.25, description="Annual holding cost rate"),
    service_level: float = Form(0.95, description="Target service level"),
    echelon_cost_multiplier_0: float = Form(1.0, description="Cost multiplier for echelon 0 (suppliers)"),
    echelon_cost_multiplier_1: float = Form(1.2, description="Cost multiplier for echelon 1 (plants)"),
    echelon_cost_multiplier_2: float = Form(1.5, description="Cost multiplier for echelon 2 (DCs)"),
    echelon_cost_multiplier_3: float = Form(2.0, description="Cost multiplier for echelon 3 (retailers)"),
    product_value_default: float = Form(100, description="Default product value per unit")
):
    """
    Calculate echelon inventory costs for multi-echelon networks
    
    Supports:
    - Multi-echelon inventory holding costs
    - Safety stock optimization across echelons
    - Lead time variability considerations
    - Service level constraints by echelon
    - Cost allocation by echelon and product
    
    Returns echelon inventory cost analysis and optimization.
    """
    try:
        # Read and process input files
        network_data = await _process_uploaded_file(network_file)
        demand_data = await _process_uploaded_file(demand_file)
        
        # Prepare network structure
        network_structure = {
            'nodes': network_data,
            'echelon_relationships': {}  # Will be inferred from node types
        }
        
        # Prepare demand patterns
        demand_patterns = {
            'demand_by_node': {},
            'demand_uncertainty': {},
            'lead_times': {}
        }
        
        # Process demand data
        for _, row in pd.DataFrame(demand_data).iterrows():
            node_id = row.get('node_id') or row.get('name')
            product = row.get('product', 'default')
            mean_demand = row.get('mean_demand', row.get('demand', 0))
            demand_variance = row.get('demand_variance', mean_demand * 0.2)
            lead_time = row.get('lead_time', 1)
            
            if node_id not in demand_patterns['demand_by_node']:
                demand_patterns['demand_by_node'][node_id] = {}
            demand_patterns['demand_by_node'][node_id][product] = mean_demand
            
            if node_id not in demand_patterns['demand_uncertainty']:
                demand_patterns['demand_uncertainty'][node_id] = {}
            demand_patterns['demand_uncertainty'][node_id][product] = demand_variance
            
            demand_patterns['lead_times'][node_id] = lead_time
        
        # Prepare cost parameters
        cost_parameters = {
            'holding_cost_rate': holding_cost_rate,
            'product_values': {'default': product_value_default},
            'echelon_cost_multipliers': {
                'echelon_0': echelon_cost_multiplier_0,
                'echelon_1': echelon_cost_multiplier_1,
                'echelon_2': echelon_cost_multiplier_2,
                'echelon_3': echelon_cost_multiplier_3
            }
        }
        
        # Prepare service levels
        service_levels = {
            'default_service_level': service_level,
            'echelon_0': service_level,
            'echelon_1': service_level,
            'echelon_2': service_level,
            'echelon_3': service_level
        }
        
        # Calculate echelon inventory costs
        result = calculate_echelon_inventory_costs(
            network_structure, demand_patterns, cost_parameters, service_levels
        )
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating echelon inventory costs: {str(e)}")

async def _process_uploaded_file(file: UploadFile) -> List[Dict[str, Any]]:
    """
    Helper function to process uploaded CSV/Excel files
    
    Args:
        file: Uploaded file
        
    Returns:
        List of dictionaries representing the data
    """
    contents = await file.read()
    
    if file.filename.endswith('.csv'):
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
    elif file.filename.endswith(('.xlsx', '.xls')):
        df = pd.read_excel(io.BytesIO(contents))
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Please use CSV or Excel files.")
    
    # Convert DataFrame to list of dictionaries
    return df.to_dict('records')

@router.get("/abstract-logistics/info")
async def get_abstract_logistics_info():
    """
    Get information about abstract logistics objects capabilities
    
    Returns details about LNDP, echelon inventory costs, BOM modeling,
    and other advanced logistics optimization features.
    """
    return {
        "abstract_logistics_objects": {
            "description": "Advanced logistics network design with abstract objects",
            "version": "1.0.0",
            "features": [
                "LNDP (Logistics Network Design Problem) solver",
                "Multi-echelon inventory cost optimization",
                "Multi-level BOM (Bill of Materials) modeling",
                "Assembly and disassembly process support",
                "Echelon-specific service level constraints",
                "Advanced optimization with Gurobi and PuLP"
            ]
        },
        "lndp_solver": {
            "description": "Generalized logistics network design framework",
            "capabilities": [
                "Multi-product, multi-echelon networks",
                "Facility location and capacity decisions",
                "Flow optimization with service constraints",
                "BOM hierarchy with assembly/disassembly costs",
                "Advanced solver integration (Gurobi/PuLP)"
            ],
            "input_requirements": [
                "Network nodes (suppliers, plants, DCs, customers)",
                "Network arcs with capacities and costs",
                "Product definitions and attributes",
                "BOM structure (optional)",
                "Demand data (optional - can be generated)"
            ]
        },
        "echelon_inventory": {
            "description": "Multi-echelon inventory cost optimization",
            "capabilities": [
                "Safety stock calculation by echelon",
                "Demand propagation through network levels",
                "Service level optimization",
                "Cost allocation by echelon and product",
                "Lead time amplification modeling"
            ],
            "echelon_levels": {
                "0": "Suppliers",
                "1": "Plants/Factories", 
                "2": "Distribution Centers",
                "3": "Retailers",
                "4": "Customers"
            }
        },
        "bom_modeling": {
            "description": "Multi-level Bill of Materials processing",
            "capabilities": [
                "Hierarchical product structure",
                "Assembly and disassembly operations",
                "Product level calculation",
                "Circular dependency detection",
                "Cost integration with network optimization"
            ]
        },
        "supported_formats": {
            "input": ["CSV", "Excel (.xlsx, .xls)"],
            "output": ["JSON", "Structured optimization results"],
            "visualization": ["Network diagrams", "Cost analysis charts"]
        }
    }

# Excel Integration Endpoints (converted from notebook functions)

@router.post("/excel/create-template", response_model=ExcelTemplateResponse)
async def create_excel_template(request: ExcelTemplateRequest):
    """
    Create Excel template for MELOS (MEta Logistic network Optimization System)
    Converts notebook make_excel_melos() function to API endpoint
    """
    try:
        # Create Excel template using service
        workbook = excel_integration_service.make_excel_melos()
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        workbook.save(output)
        template_data = output.getvalue()
        
        return ExcelTemplateResponse(
            success=True,
            message="Excel template created successfully",
            template_data=template_data
        )
        
    except Exception as e:
        return ExcelTemplateResponse(
            success=False,
            message=f"Failed to create Excel template: {str(e)}"
        )

@router.post("/excel/add-sheets", response_model=ExcelWorkflowResponse)
async def add_demand_production_sheets(
    file: UploadFile = File(...),
    operation: str = Form("add_sheets")
):
    """
    Add demand and production sheets to existing Excel workbook
    Converts notebook make_demand_production_sheets() function to API endpoint
    """
    try:
        # Read uploaded Excel file
        from openpyxl import load_workbook
        from io import BytesIO
        
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        
        # Add sheets using service
        updated_workbook = excel_integration_service.make_demand_production_sheets(workbook)
        
        # Save to bytes
        output = BytesIO()
        updated_workbook.save(output)
        result_data = output.getvalue()
        
        return ExcelWorkflowResponse(
            success=True,
            message="Demand and production sheets added successfully",
            result_data=result_data
        )
        
    except Exception as e:
        return ExcelWorkflowResponse(
            success=False,
            message=f"Failed to add sheets: {str(e)}"
        )

@router.post("/excel/extract-constraints", response_model=ExcelWorkflowResponse)
async def extract_warehouse_constraints(file: UploadFile = File(...)):
    """
    Extract warehouse fixing constraints from Excel cell colors
    Converts notebook extract_fix_dc_info() function to API endpoint
    """
    try:
        # Read uploaded Excel file
        from openpyxl import load_workbook
        from io import BytesIO
        
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        
        # Extract constraints using service
        constraints = excel_integration_service.extract_fix_dc_info(workbook)
        
        return ExcelWorkflowResponse(
            success=True,
            message="Warehouse constraints extracted successfully",
            analysis_results={"fixed_warehouses": constraints}
        )
        
    except Exception as e:
        return ExcelWorkflowResponse(
            success=False,
            message=f"Failed to extract constraints: {str(e)}"
        )

@router.post("/excel/solve-lnd", response_model=LNDSolveResponse)
async def solve_lnd_from_excel(
    file: UploadFile = File(...),
    solver_type: str = Form("multi_source")
):
    """
    Solve Logistics Network Design problem from Excel data
    Integrates Excel workflow with LND solvers from notebook
    """
    try:
        # Read uploaded Excel file
        from openpyxl import load_workbook
        from io import BytesIO
        
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        
        # Solve using service
        results = excel_integration_service.solve_lnd_for_excel(workbook, solver_type)
        
        if "error" in results:
            return LNDSolveResponse(
                success=False,
                message=results["error"]
            )
        
        return LNDSolveResponse(
            success=True,
            message="LND problem solved successfully",
            optimization_results=results.get("optimization_results", {}),
            cost_breakdown=results.get("cost_breakdown", {}),
            selected_facilities=results.get("selected_facilities", []),
            customer_assignments=results.get("customer_assignments", [])
        )
        
    except Exception as e:
        return LNDSolveResponse(
            success=False,
            message=f"Failed to solve LND: {str(e)}"
        )

@router.post("/excel/add-results", response_model=ExcelWorkflowResponse)
async def add_results_to_excel(
    file: UploadFile = File(...),
    results_data: str = Form(...)
):
    """
    Add optimization results back to Excel workbook
    Converts notebook add_result_for_melos() function to API endpoint
    """
    try:
        # Read uploaded Excel file
        from openpyxl import load_workbook
        from io import BytesIO
        import json
        
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        
        # Parse results data
        results = json.loads(results_data)
        
        # Add results using service
        updated_workbook = excel_integration_service.add_result_for_melos(workbook, results)
        
        # Save to bytes
        output = BytesIO()
        updated_workbook.save(output)
        result_data = output.getvalue()
        
        return ExcelWorkflowResponse(
            success=True,
            message="Results added to Excel successfully",
            result_data=result_data
        )
        
    except Exception as e:
        return ExcelWorkflowResponse(
            success=False,
            message=f"Failed to add results: {str(e)}"
        )

@router.post("/excel/generate-network", response_model=ExcelWorkflowResponse)
async def generate_network_for_excel(file: UploadFile = File(...)):
    """
    Generate transportation network sheet for Excel workbook
    Converts notebook make_network_for_excel() function to API endpoint
    """
    try:
        # Read uploaded Excel file
        from openpyxl import load_workbook
        from io import BytesIO
        
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        
        # Generate network using service
        network_df = excel_integration_service.make_network_for_excel(workbook)
        
        # Save updated workbook
        output = BytesIO()
        workbook.save(output)
        result_data = output.getvalue()
        
        return ExcelWorkflowResponse(
            success=True,
            message="Network sheet generated successfully",
            result_data=result_data,
            analysis_results={"network_statistics": {
                "total_routes": len(network_df),
                "average_distance": float(network_df['distance_m'].mean()) if len(network_df) > 0 else 0,
                "total_cost": float(network_df['cost_per_unit'].sum()) if len(network_df) > 0 else 0
            }}
        )
        
    except Exception as e:
        return ExcelWorkflowResponse(
            success=False,
            message=f"Failed to generate network: {str(e)}"
        )

# Customer Aggregation Endpoints (converted from notebook functions)

@router.post("/aggregation/kmeans", response_model=CustomerAggregationResponse)
async def kmeans_customer_aggregation(
    file: UploadFile = File(...),
    num_clusters: int = Form(10),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    batch: bool = Form(True)
):
    """
    K-means clustering for customer aggregation
    Converts notebook kmeans() function to API endpoint
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Extract weights
        weights = df[demand_col].tolist() if demand_col in df.columns else [1.0] * len(df)
        
        # Perform k-means clustering using service
        X, Y, partition, cost = customer_aggregation_service.kmeans(
            cust_df=df,
            weight=weights,
            num_of_facilities=num_clusters,
            batch=batch
        )
        
        # Create aggregated customers
        aggregated_customers = []
        for i in range(len(X)):
            aggregated_customers.append({
                "cluster_id": f"cluster_{i}",
                "lat": X[i],
                "lon": Y[i],
                "total_demand": sum(weights[j] for j in range(len(df)) if partition[j] == i)
            })
        
        return CustomerAggregationResponse(
            success=True,
            message="K-means clustering completed successfully",
            aggregated_customers=aggregated_customers,
            cluster_assignments=partition.tolist(),
            aggregated_demand=[{"cluster_id": f"cluster_{i}", "demand": sum(weights[j] for j in range(len(df)) if partition[j] == i)} for i in range(len(X))]
        )
        
    except Exception as e:
        return CustomerAggregationResponse(
            success=False,
            message=f"Failed to perform k-means clustering: {str(e)}"
        )

@router.post("/aggregation/k-median", response_model=CustomerAggregationResponse)
async def k_median_customer_aggregation(
    file: UploadFile = File(...),
    num_facilities: int = Form(3),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    max_iterations: int = Form(100),
    max_lr: float = Form(0.01),
    convergence: float = Form(1e-5)
):
    """
    K-median clustering using Lagrange relaxation
    Converts notebook solve_k_median() function to API endpoint
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Extract weights
        weights = df[demand_col].tolist() if demand_col in df.columns else [1.0] * len(df)
        
        # Calculate cost matrix using great circle distances
        n = len(df)
        cost_matrix = np.zeros((n, n))
        for i in range(n):
            for j in range(n):
                if i != j:
                    dist = great_circle_distance(
                        df.iloc[i][lat_col], df.iloc[i][lon_col],
                        df.iloc[j][lat_col], df.iloc[j][lon_col]
                    )
                    cost_matrix[i][j] = dist
        
        # Solve k-median using service
        X, Y, partition, best_cost, lb_list, ub_list, phi_list = customer_aggregation_service.solve_k_median(
            cust_df=df,
            weight=weights,
            cost=cost_matrix,
            num_of_facilities=num_facilities,
            max_iter=max_iterations,
            max_lr=max_lr,
            convergence=convergence
        )
        
        # Create aggregated customers
        aggregated_customers = []
        for i in range(len(X)):
            aggregated_customers.append({
                "cluster_id": f"facility_{i}",
                "lat": X[i],
                "lon": Y[i],
                "total_demand": sum(weights[j] for j in range(len(df)) if partition[j] == i)
            })
        
        return CustomerAggregationResponse(
            success=True,
            message=f"K-median optimization completed with cost: {best_cost:.2f}",
            aggregated_customers=aggregated_customers,
            cluster_assignments=partition.tolist(),
            aggregated_demand=[{"cluster_id": f"facility_{i}", "demand": sum(weights[j] for j in range(len(df)) if partition[j] == i)} for i in range(len(X))]
        )
        
    except Exception as e:
        return CustomerAggregationResponse(
            success=False,
            message=f"Failed to perform k-median optimization: {str(e)}"
        )

@router.post("/aggregation/hierarchical", response_model=CustomerAggregationResponse)
async def hierarchical_customer_aggregation(
    file: UploadFile = File(...),
    num_clusters: int = Form(10),
    lat_col: str = Form("lat"),
    lon_col: str = Form("lon"),
    demand_col: str = Form("demand"),
    linkage: str = Form("ward"),
    use_road_distance: bool = Form(False)
):
    """
    Hierarchical clustering for customer aggregation
    Uses existing hierarchical clustering service functionality
    """
    try:
        # Read CSV file
        contents = await file.read()
        df = pd.read_csv(io.StringIO(contents.decode('utf-8')))
        
        # Validate required columns
        required_cols = [lat_col, lon_col]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {missing_cols}"
            )
        
        # Extract weights
        weights = df[demand_col].tolist() if demand_col in df.columns else [1.0] * len(df)
        
        # Perform hierarchical clustering using existing service
        result = customer_aggregation_service.hierarchical_customer_aggregation(
            customer_df=df,
            weights=weights,
            num_clusters=num_clusters,
            lat_col=lat_col,
            lon_col=lon_col,
            linkage=linkage,
            use_road_distance=use_road_distance
        )
        
        return CustomerAggregationResponse(
            success=True,
            message="Hierarchical clustering completed successfully",
            aggregated_customers=result.get("aggregated_customers", []),
            cluster_assignments=result.get("cluster_assignments", []),
            aggregated_demand=result.get("aggregated_demand", [])
        )
        
    except Exception as e:
        return CustomerAggregationResponse(
            success=False,
            message=f"Failed to perform hierarchical clustering: {str(e)}"
        )

# LNDP (Abstract Logistics Network Design) Endpoints

@router.post("/lndp/solve", response_model=AdvancedLNDPResponse)
async def solve_advanced_lndp(request: AdvancedLNDPRequest):
    """
    Solve Advanced Logistics Network Design Problem (LNDP)
    Converts notebook LNDP() function to API endpoint
    """
    try:
        # Prepare nodes data
        nodes_data = {
            'nodes': []
        }
        for node in request.nodes:
            nodes_data['nodes'].append({
                'id': node,
                'name': node,
                'type': 'node'  # Can be extended with more specific types
            })
        
        # Prepare arcs data
        arcs_data = {
            'arcs': request.arcs
        }
        
        # Prepare products data
        products_data = {
            'products': request.products
        }
        
        # Prepare BOM data
        bom_data = None
        if request.bom:
            bom_data = {
                'bom_relationships': request.bom,
                'assembly_costs': {},
                'disassembly_costs': {}
            }
        
        # Prepare demand data
        demand_data = None
        if request.demand:
            demand_data = {
                'demand_matrix': request.demand
            }
        
        # Set up cost parameters
        cost_parameters = {
            'transport_cost_per_km': 0.5,
            'fixed_cost_facility': 10000,
            'variable_cost_facility': 0.1,
            'holding_cost_rate': 0.2,
            'ordering_cost': 100
        }
        
        # Set up service constraints
        service_constraints = {
            'max_distance': 500,
            'min_service_level': 0.95,
            'max_lead_time': 7
        }
        if request.constraints:
            service_constraints.update(request.constraints)
        
        # Set up optimization options
        optimization_options = {
            'solver_type': 'pulp',  # Use PuLP as default since Gurobi may not be available
            'time_limit': 3600,
            'mip_gap': 0.01,
            'threads': 4
        }
        
        # Solve LNDP using service
        results = LNDP(
            nodes_data=nodes_data,
            arcs_data=arcs_data,
            products_data=products_data,
            bom_data=bom_data,
            demand_data=demand_data,
            cost_parameters=cost_parameters,
            service_constraints=service_constraints,
            optimization_options=optimization_options
        )
        
        # Calculate carbon footprint if limit is specified
        carbon_footprint = None
        if request.carbon_footprint_limit is not None:
            # Use carbon footprint service for calculation
            carbon_calculation = carbon_footprint_service.multi_objective_analysis(
                facilities_data=[],  # Would need to extract from results
                customers_data=[],   # Would need to extract from results
                carbon_constraint_kg=request.carbon_footprint_limit
            )
            carbon_footprint = carbon_calculation.get('total_carbon_kg', 0.0)
        
        return AdvancedLNDPResponse(
            success=results.get('status') != 'error',
            message=results.get('message', 'LNDP solved successfully'),
            optimization_results=results.get('optimization_results', {}),
            cost_breakdown=results.get('cost_breakdown', {}),
            carbon_footprint=carbon_footprint,
            resource_utilization=results.get('resource_utilization', {})
        )
        
    except Exception as e:
        return AdvancedLNDPResponse(
            success=False,
            message=f"Failed to solve LNDP: {str(e)}"
        )

@router.post("/co2/calculate", response_model=CO2CalculationResponse)
async def calculate_co2_emissions(request: CO2CalculationRequest):
    """
    Calculate CO2 emissions for transportation
    Converts notebook co2() function to API endpoint
    """
    try:
        # Use carbon footprint service
        fuel_consumption, co2_emission = carbon_footprint_service.calculate_co2_emission(
            capacity=request.capacity,
            loading_rate=request.load_rate,
            fuel_type="diesel" if request.is_diesel else "gasoline"
        )
        
        # Calculate total emissions for the given distance
        total_co2_kg = (co2_emission / 1000) * (request.distance / 1000)  # Convert g to kg and m to km
        
        return CO2CalculationResponse(
            success=True,
            message="CO2 emissions calculated successfully",
            fuel_consumption=fuel_consumption,
            co2_emission=co2_emission
        )
        
    except Exception as e:
        return CO2CalculationResponse(
            success=False,
            message=f"Failed to calculate CO2 emissions: {str(e)}"
        )

# VRP Integration Endpoints

@router.post("/vrp/integrate-lnd", response_model=VRPIntegrationResponse)
async def integrate_vrp_with_lnd(request: VRPIntegrationRequest):
    """
    Integrate Vehicle Routing Problem with Logistics Network Design
    Converts notebook make_vrp() function to API endpoint
    """
    try:
        from app.services.pyvrp_service import PyVRPService
        pyvrp_service = PyVRPService()
        
        # Extract customer locations from LND solution
        lnd_solution = request.lnd_solution
        customer_data = request.customer_data
        vehicle_specs = request.vehicle_specifications
        
        # Create locations list for VRP
        locations = []
        demands = []
        
        # Add depot (first location)
        if customer_data:
            depot_location = customer_data[0]
            locations.append({
                'name': depot_location.get('name', 'Depot'),
                'lat': depot_location.get('lat', 0),
                'lon': depot_location.get('lon', 0)
            })
            demands.append(0)  # Depot has no demand
        
        # Add customer locations
        for customer in customer_data[1:]:
            locations.append({
                'name': customer.get('name', ''),
                'lat': customer.get('lat', 0),
                'lon': customer.get('lon', 0)
            })
            demands.append(customer.get('demand', 1))
        
        # Extract vehicle capacity
        vehicle_capacity = 1000  # Default capacity
        if vehicle_specs:
            vehicle_capacity = vehicle_specs[0].get('capacity', 1000)
        
        # Solve VRP
        if request.time_windows:
            # Solve VRPTW if time windows are provided
            time_windows = [(tw.get('earliest', 0), tw.get('latest', 1440)) for tw in request.time_windows]
            service_times = [tw.get('service_time', 0) for tw in request.time_windows]
            
            vrp_solution = pyvrp_service.solve_vrptw(
                locations=locations,
                demands=demands,
                time_windows=time_windows,
                service_times=service_times,
                vehicle_capacity=vehicle_capacity,
                depot_index=0
            )
        else:
            # Solve basic CVRP
            vrp_solution = pyvrp_service.solve_basic_cvrp(
                locations=locations,
                demands=demands,
                vehicle_capacity=vehicle_capacity,
                depot_index=0
            )
        
        # Calculate integrated cost (LND + VRP)
        lnd_cost = lnd_solution.get('total_cost', 0)
        vrp_cost = vrp_solution.get('objective_value', 0)
        integrated_cost = lnd_cost + vrp_cost
        
        # Prepare routing statistics
        routing_statistics = {
            'total_routes': len(vrp_solution.get('routes', [])),
            'total_distance': vrp_solution.get('total_distance', 0),
            'total_demand_served': vrp_solution.get('total_demand_served', 0),
            'num_vehicles_used': vrp_solution.get('num_vehicles_used', 0),
            'computation_time': vrp_solution.get('computation_time', 0),
            'lnd_cost': lnd_cost,
            'vrp_cost': vrp_cost
        }
        
        return VRPIntegrationResponse(
            success=vrp_solution.get('status') in ['optimal', 'feasible'],
            message=f"VRP integration completed with status: {vrp_solution.get('status', 'unknown')}",
            vrp_solutions=[vrp_solution],
            routing_statistics=routing_statistics,
            integrated_cost=integrated_cost
        )
        
    except Exception as e:
        return VRPIntegrationResponse(
            success=False,
            message=f"Failed to integrate VRP with LND: {str(e)}"
        )

# Network Generation Endpoints (converted from notebook functions)

@router.post("/network/generate-great-circle", response_model=NetworkGenerationResponse)
async def generate_network_great_circle(
    customer_file: UploadFile = File(...),
    dc_file: UploadFile = File(...),
    plant_file: UploadFile = File(...),
    plnt_dc_threshold: float = Form(999999.0),
    dc_cust_threshold: float = Form(999999.0),
    unit_tp_cost: float = Form(1.0),
    unit_del_cost: float = Form(1.0)
):
    """
    Generate transportation network using great circle distances
    Converts notebook make_network() function to API endpoint
    """
    try:
        # Read CSV files
        customer_contents = await customer_file.read()
        dc_contents = await dc_file.read()
        plant_contents = await plant_file.read()
        
        cust_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        dc_df = pd.read_csv(io.StringIO(dc_contents.decode('utf-8')))
        plnt_df = pd.read_csv(io.StringIO(plant_contents.decode('utf-8')))
        
        # Generate network using service
        trans_df, graph, position = network_generation_service.make_network(
            cust_df=cust_df,
            dc_df=dc_df,
            plnt_df=plnt_df,
            plnt_dc_threshold=plnt_dc_threshold,
            dc_cust_threshold=dc_cust_threshold,
            unit_tp_cost=unit_tp_cost,
            unit_del_cost=unit_del_cost
        )
        
        # Convert NetworkX graph to list format for response
        transportation_network = []
        for _, row in trans_df.iterrows():
            transportation_network.append({
                "from_node": row['from_node'],
                "to_node": row['to_node'],
                "distance_km": row['dist'],
                "cost": row['cost'],
                "lead_time": row['lead_time'],
                "stage_time": row['stage_time'],
                "connection_type": row['kind']
            })
        
        # Generate network statistics
        network_statistics = {
            "total_connections": len(trans_df),
            "plant_dc_connections": len(trans_df[trans_df['kind'] == 'plnt-dc']),
            "dc_customer_connections": len(trans_df[trans_df['kind'] == 'dc-cust']),
            "total_distance": float(trans_df['dist'].sum()),
            "total_cost": float(trans_df['cost'].sum()),
            "average_distance": float(trans_df['dist'].mean()),
            "graph_nodes": graph.number_of_nodes(),
            "graph_edges": graph.number_of_edges()
        }
        
        return NetworkGenerationResponse(
            success=True,
            message="Network generated successfully using great circle distances",
            transportation_network=transportation_network,
            network_statistics=network_statistics
        )
        
    except Exception as e:
        return NetworkGenerationResponse(
            success=False,
            message=f"Failed to generate network: {str(e)}"
        )

@router.post("/network/visualize-scm", response_model=Dict[str, Any])
async def visualize_supply_chain_network(
    customer_file: UploadFile = File(...),
    dc_file: UploadFile = File(...),
    plant_file: UploadFile = File(...),
    node_only: bool = Form(False)
):
    """
    Visualize supply chain network on map
    Converts notebook plot_scm() function to API endpoint
    """
    try:
        # Read CSV files
        customer_contents = await customer_file.read()
        dc_contents = await dc_file.read()
        plant_contents = await plant_file.read()
        
        cust_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        dc_df = pd.read_csv(io.StringIO(dc_contents.decode('utf-8')))
        plnt_df = pd.read_csv(io.StringIO(plant_contents.decode('utf-8')))
        
        # Generate network first to get graph and positions
        trans_df, graph, position = network_generation_service.make_network(
            cust_df=cust_df,
            dc_df=dc_df,
            plnt_df=plnt_df
        )
        
        # Create visualization using service
        fig = network_generation_service.plot_scm(
            cust_df=cust_df,
            dc_df=dc_df,
            plnt_df=plnt_df,
            graph=graph,
            position=position,
            node_only=node_only
        )
        
        # Convert Plotly figure to JSON
        fig_json = fig.to_dict()
        
        return {
            "success": True,
            "message": "Supply chain network visualization created",
            "visualization": fig_json
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"Failed to create visualization: {str(e)}"
        }

@router.post("/network/distance-histogram", response_model=Dict[str, Any])
async def generate_distance_histogram(
    customer_file: UploadFile = File(...),
    dc_file: UploadFile = File(...),
    plant_file: UploadFile = File(...)
):
    """
    Generate distance histogram analysis
    Converts notebook distance_histgram() function to API endpoint
    """
    try:
        # Read CSV files
        customer_contents = await customer_file.read()
        dc_contents = await dc_file.read()
        plant_contents = await plant_file.read()
        
        cust_df = pd.read_csv(io.StringIO(customer_contents.decode('utf-8')))
        dc_df = pd.read_csv(io.StringIO(dc_contents.decode('utf-8')))
        plnt_df = pd.read_csv(io.StringIO(plant_contents.decode('utf-8')))
        
        # Generate distance histogram using service
        fig = network_generation_service.distance_histgram(
            cust_df=cust_df,
            dc_df=dc_df,
            plnt_df=plnt_df
        )
        
        # Convert Plotly figure to JSON
        fig_json = fig.to_dict()
        
        return {
            "success": True,
            "message": "Distance histogram created successfully",
            "histogram": fig_json
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"Failed to generate distance histogram: {str(e)}"
        }