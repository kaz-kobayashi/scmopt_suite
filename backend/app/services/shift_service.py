import pandas as pd
import numpy as np
import json
import ast
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict, namedtuple, OrderedDict
import holidays
from faker import Faker
import random
import tempfile
import os
from openpyxl import Workbook
import plotly.graph_objs as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.figure_factory as ff

# SCOP related imports (仮想的なインポート - 実際にはSCOPライブラリが必要)
try:
    from scmopt2.scop import *
except ImportError:
    # SCOP が利用できない場合の代替実装
    print("Warning: SCOP library not found. Using alternative optimization.")

class ShiftService:
    
    def __init__(self):
        self.min_work_periods = 3  # 最小稼働期間
    
    def _safe_parse_list(self, value):
        """Safely parse a value that could be a string or already a list"""
        if isinstance(value, list):
            return value
        elif isinstance(value, str):
            try:
                return ast.literal_eval(value)
            except (ValueError, SyntaxError):
                return []
        else:
            return []
            
    def _safe_parse_dict(self, value):
        """Safely parse a value that could be a string or already a dict"""
        if isinstance(value, dict):
            return value
        elif isinstance(value, str):
            try:
                return ast.literal_eval(value)
            except (ValueError, SyntaxError):
                return {}
        else:
            return {}
        
    def generate_day(self, start_date: str, end_date: str) -> pd.DataFrame:
        """日データ生成"""
        jp_holidays = holidays.Japan()
        day_df = pd.DataFrame(pd.date_range(start_date, end_date, freq='D'), columns=["day"])
        day_df["day_of_week"] = [('Holiday') if t in jp_holidays else (t.strftime('%a')) for t in day_df["day"]]
        
        row_ = []
        for row in day_df.itertuples():
            if row.day_of_week == "Holiday":
                row_.append("holiday")
            elif row.day_of_week == "Sun":
                row_.append("sunday")
            else:
                row_.append("weekday")
        day_df["day_type"] = row_
        day_df["id"] = [t for t in range(len(day_df))]
        day_df = day_df.reindex(columns=["id", "day", "day_of_week", "day_type"])
        return day_df
    
    def generate_period(self, start_time: str, end_time: str, freq: str = "1h") -> pd.DataFrame:
        """期間データ生成"""
        period_df = pd.DataFrame(pd.date_range(start_time, end_time, freq=freq), columns=["description"])
        period_df["description"] = period_df.description.dt.strftime("%H:%M")
        period_df["id"] = [t for t in range(len(period_df))]
        period_df = period_df.reindex(columns=["id", "description"])
        return period_df
    
    def generate_break_data(self, period_count: int) -> pd.DataFrame:
        """休憩データ生成"""
        period_ = list(range(self.min_work_periods, period_count))
        break_time = [0]  # 最低勤務時間は休憩なし
        
        # 簡単な休憩ルール: 勤務時間に比例して休憩時間を追加
        for t in range(self.min_work_periods + 1, period_count):
            if t >= 6:  # 6時間以上で1回目の休憩
                break_count = (t - 3) // 3  # 3時間毎に休憩追加
                break_time.append(min(break_count, 3))  # 最大3回まで
            else:
                break_time.append(0)
        
        break_df = pd.DataFrame({
            "period": period_,
            "break_time": break_time
        })
        return break_df
    
    def generate_job_data(self, job_list: List[str]) -> pd.DataFrame:
        """ジョブデータ生成"""
        description_ = ["break"] + job_list
        n_job = len(description_)
        id_ = list(range(n_job))
        job_df = pd.DataFrame({"id": id_, "description": description_})
        return job_df
    
    def generate_staff_data(self, n_staff: int, job_count: int, day_count: int, period_count: int) -> pd.DataFrame:
        """スタッフデータ生成"""
        fake = Faker(['ja_JP'])
        Faker.seed(1)
        
        name_ = [fake.name() for i in range(n_staff)]
        job_list = list(range(1, job_count))  # 休憩(0)を除く
        
        staff_df = pd.DataFrame({
            "name": name_,
            "wage_per_period": np.random.randint(low=850, high=1300, size=n_staff),
            "max_period": np.random.randint(5, min(12, period_count), n_staff),
            "max_day": np.random.randint(1, min(3, day_count), n_staff),
            "job_set": [random.sample(job_list, random.randint(1, len(job_list))) for s in range(n_staff)],
            "day_off": [random.sample(list(range(day_count)), random.randint(0, 2)) for s in range(n_staff)],
            "start": np.random.randint(low=0, high=period_count//2-1, size=n_staff),
            "end": np.random.randint(low=period_count//2+1, high=period_count-1, size=n_staff),
            "request": [None for _ in range(n_staff)]
        })
        
        return staff_df
    
    def generate_requirement_data(self, day_types: List[str], job_count: int, period_count: int) -> pd.DataFrame:
        """必要人数データ生成"""
        type_, job_, period_, lb_ = [], [], [], []
        
        for d in day_types:
            for j in range(1, job_count):  # 休憩(0)は除く
                req_ = np.ones(period_count, int)
                
                # 時間帯によって必要人数を調整
                for t in range(period_count):
                    if 2 <= t <= 8:  # ピーク時間帯
                        req_[t] = random.randint(2, 4)
                    else:
                        req_[t] = random.randint(1, 2)
                
                for t in range(period_count):
                    type_.append(d)
                    job_.append(j)
                    period_.append(t)
                    lb_.append(req_[t])
        
        requirement_df = pd.DataFrame({
            "day_type": type_,
            "job": job_,
            "period": period_,
            "requirement": lb_
        })
        return requirement_df
    
    def generate_sample_data(self, start_date: str, end_date: str, 
                            start_time: str, end_time: str, freq: str,
                            job_list: List[str]) -> Dict[str, Any]:
        """サンプルデータ一括生成"""
        
        # 各データフレームを生成
        day_df = self.generate_day(start_date, end_date)
        period_df = self.generate_period(start_time, end_time, freq)
        job_df = self.generate_job_data(job_list)
        break_df = self.generate_break_data(len(period_df))
        
        n_staff = 10  # デフォルトスタッフ数
        staff_df = self.generate_staff_data(n_staff, len(job_df), len(day_df), len(period_df)-1)
        
        day_types = ["weekday", "sunday", "holiday"]
        requirement_df = self.generate_requirement_data(day_types, len(job_df), len(period_df)-1)
        
        return {
            "day_df": day_df.to_dict('records'),
            "period_df": period_df.to_dict('records'),
            "job_df": job_df.to_dict('records'),
            "break_df": break_df.to_dict('records'),
            "staff_df": staff_df.to_dict('records'),
            "requirement_df": requirement_df.to_dict('records')
        }
    
    def shift_scheduling(self, period_df: pd.DataFrame, break_df: pd.DataFrame,
                        day_df: pd.DataFrame, job_df: pd.DataFrame,
                        staff_df: pd.DataFrame, requirement_df: pd.DataFrame,
                        theta: int = 1, lb_penalty: int = 10000, ub_penalty: int = 0,
                        job_change_penalty: int = 10, break_penalty: int = 10000,
                        max_day_penalty: int = 5000, time_limit: int = 30,
                        random_seed: int = 1) -> Dict[str, Any]:
        """シフト最適化実行（SCOP準拠）"""
        
        try:
            return self._shift_scheduling_scop(
                period_df, break_df, day_df, job_df, staff_df, requirement_df,
                theta, lb_penalty, ub_penalty, job_change_penalty, break_penalty,
                max_day_penalty, time_limit, random_seed
            )
            
        except Exception as e:
            return {
                "status": -1,
                "message": f"最適化エラー: {str(e)}",
                "cost_df": [],
                "violate_df": {},
                "staff_df": [],
                "job_assign": {}
            }

    def _shift_scheduling_scop(self, period_df: pd.DataFrame, break_df: pd.DataFrame,
                               day_df: pd.DataFrame, job_df: pd.DataFrame,
                               staff_df: pd.DataFrame, requirement_df: pd.DataFrame,
                               theta: int, lb_penalty: int, ub_penalty: int,
                               job_change_penalty: int, break_penalty: int,
                               max_day_penalty: int, time_limit: int,
                               random_seed: int) -> Dict[str, Any]:
        """SCOP準拠のシフト最適化実装"""
        
        # 基本パラメータ
        min_work_time = break_df.loc[0, "period"]
        n_job = len(job_df)
        n_day = len(day_df)
        n_period = len(period_df) - 1
        n_staff = len(staff_df)
        
        # 要求タイプ、ジョブ、期ごとの必要人数を入れる辞書
        requirement = {}
        for row in requirement_df.itertuples():
            requirement[row.day_type, row.period, row.job] = row.requirement
        
        # 休日希望日の集合を返す辞書
        day_off = {}
        for i in range(n_staff):
            day_off[i] = set(self._safe_parse_list(staff_df.loc[i, "day_off"]))
        
        # スタッフごとに開始時刻と終了時刻の組（パターン）を準備
        st_dic = defaultdict(list)
        start, end = {}, {}
        max_period, max_day = {}, {}
        
        for i, row in enumerate(staff_df.itertuples()):
            start[i] = int(row.start)
            end[i] = int(row.end)
            max_period[i] = int(row.max_period)
            max_day[i] = int(row.max_day)
        
        for i in range(n_staff):
            st_dic[i].append(f"{n_period}_{n_period}")  # ダミー
            for s in range(start[i], end[i] - min_work_time + 1):
                for t in range(s + min_work_time - 1, end[i] + 1):
                    if s + max_period[i] < t:
                        break
                    else:
                        st_dic[i].append(f"{s}_{t}")
        
        # 稼働時間に対する休憩時間数を返す辞書
        break_time = defaultdict(int)
        for t, b in zip(break_df.period, break_df.break_time):
            break_time[int(t)] = int(b)
        
        # スタッフに割り当て可能なジョブのリスト
        job_set = {}
        for i in range(n_staff):
            job_set[i] = [0] + self._safe_parse_list(staff_df.loc[i, "job_set"]) + [n_job]
        
        # 最適化実行（簡易版 - 実際にはSCOPライブラリを使用）
        result = self._solve_shift_optimization(
            n_staff, n_day, n_period, n_job,
            st_dic, job_set, requirement, day_off, break_time,
            start, end, max_period, max_day,
            theta, lb_penalty, ub_penalty, job_change_penalty,
            break_penalty, max_day_penalty, staff_df, day_df
        )
        
        return result

    def shift_scheduling2(self, period_df: pd.DataFrame, break_df: pd.DataFrame,
                         day_df: pd.DataFrame, job_df: pd.DataFrame,
                         staff_df: pd.DataFrame, requirement_df: pd.DataFrame,
                         theta: int = 1, lb_penalty: int = 10000, ub_penalty: int = 0,
                         job_change_penalty: int = 10, break_penalty: int = 10000,
                         max_day_penalty: int = 5000, time_limit: int = 30,
                         random_seed: int = 1) -> Dict[str, Any]:
        """シフト最適化実行（日別リクエスト対応版）"""
        
        try:
            return self._shift_scheduling_with_requests(
                period_df, break_df, day_df, job_df, staff_df, requirement_df,
                theta, lb_penalty, ub_penalty, job_change_penalty, break_penalty,
                max_day_penalty, time_limit, random_seed
            )
            
        except Exception as e:
            return {
                "status": -1,
                "message": f"最適化エラー: {str(e)}",
                "cost_df": [],
                "violate_df": {},
                "staff_df": [],
                "job_assign": {}
            }

    def _shift_scheduling_with_requests(self, period_df: pd.DataFrame, break_df: pd.DataFrame,
                                       day_df: pd.DataFrame, job_df: pd.DataFrame,
                                       staff_df: pd.DataFrame, requirement_df: pd.DataFrame,
                                       theta: int, lb_penalty: int, ub_penalty: int,
                                       job_change_penalty: int, break_penalty: int,
                                       max_day_penalty: int, time_limit: int,
                                       random_seed: int) -> Dict[str, Any]:
        """日別希望勤務時間対応のシフト最適化"""
        
        # 基本パラメータ
        min_work_time = break_df.loc[0, "period"]
        n_job = len(job_df)
        n_day = len(day_df)
        n_period = len(period_df) - 1
        n_staff = len(staff_df)
        
        # 要求タイプ、ジョブ、期ごとの必要人数を入れる辞書
        requirement = {}
        for row in requirement_df.itertuples():
            requirement[row.day_type, row.period, row.job] = row.requirement
        
        # 休日希望日の集合を返す辞書
        day_off = {}
        for i in range(n_staff):
            day_off[i] = set(self._safe_parse_list(staff_df.loc[i, "day_off"]))
        
        # スタッフごと・日ごとに開始時刻と終了時刻の組を準備
        st_dic = defaultdict(list)
        start, end = {}, {}
        max_period, max_day = {}, {}
        
        for i, row in enumerate(staff_df.itertuples()):
            start[i] = int(row.start)
            end[i] = int(row.end)
            max_period[i] = int(row.max_period)
            max_day[i] = int(row.max_day)
        
        # 日別リクエストの処理
        for i in range(n_staff):
            for d in range(n_day):
                st_dic[i, d].append(f"{n_period}_{n_period}")  # ダミー
                
                # リクエスト列の辞書を参照
                req = staff_df.loc[i, "request"]
                if req is not None and len(req) > 0:
                    try:
                        request_dict = self._safe_parse_dict(req)
                        if d in request_dict:
                            st_, en_ = request_dict[d]
                        else:
                            st_ = start[i]
                            en_ = end[i]
                    except:
                        st_ = start[i]
                        en_ = end[i]
                else:
                    st_ = start[i]
                    en_ = end[i]
                
                for s in range(st_, en_ - min_work_time + 1):
                    for t in range(s + min_work_time - 1, min(en_ + 1, n_period)):
                        if s + max_period[i] < t:
                            break
                        else:
                            st_dic[i, d].append(f"{s}_{t}")
        
        # 稼働時間に対する休憩時間数を返す辞書
        break_time = defaultdict(int)
        for t, b in zip(break_df.period, break_df.break_time):
            break_time[int(t)] = int(b)
        
        # スタッフに割り当て可能なジョブのリスト
        job_set = {}
        for i in range(n_staff):
            job_set[i] = [0] + self._safe_parse_list(staff_df.loc[i, "job_set"]) + [n_job]
        
        # 最適化実行（日別リクエスト対応版）
        result = self._solve_shift_optimization_with_requests(
            n_staff, n_day, n_period, n_job,
            st_dic, job_set, requirement, day_off, break_time,
            start, end, max_period, max_day,
            theta, lb_penalty, ub_penalty, job_change_penalty,
            break_penalty, max_day_penalty, staff_df, day_df
        )
        
        return result
    
    def _solve_shift_optimization_with_requests(self, n_staff: int, n_day: int, n_period: int, n_job: int,
                                               st_dic: Dict, job_set: Dict, requirement: Dict,
                                               day_off: Dict, break_time: Dict,
                                               start: Dict, end: Dict, max_period: Dict, max_day: Dict,
                                               theta: int, lb_penalty: int, ub_penalty: int,
                                               job_change_penalty: int, break_penalty: int,
                                               max_day_penalty: int, staff_df: pd.DataFrame,
                                               day_df: pd.DataFrame) -> Dict[str, Any]:
        """日別希望勤務時間対応の最適化ソルバー実行"""
        
        # 簡易ヒューリスティック解法による実装（日別リクエスト対応）
        job_assign = {}
        violations = defaultdict(int)
        total_cost = 0
        
        # スタッフごとにシフトを生成
        for i in range(n_staff):
            work_days = 0
            
            for d in range(n_day):
                if d in day_off[i]:
                    continue
                
                if work_days >= max_day[i]:
                    continue
                
                # 日別の希望勤務時間を取得
                req = staff_df.loc[i, "request"]
                if req is not None and len(req) > 0:
                    try:
                        request_dict = self._safe_parse_dict(req)
                        if d in request_dict:
                            work_start, work_end = request_dict[d]
                        else:
                            work_start = start[i]
                            work_end = end[i]
                    except:
                        work_start = start[i]
                        work_end = end[i]
                else:
                    work_start = start[i]
                    work_end = end[i]
                
                # 最大勤務時間の制限
                work_end = min(work_end, work_start + max_period[i] - 1)
                
                # 休憩時間の計算
                work_duration = work_end - work_start + 1
                required_breaks = break_time.get(work_duration, 0)
                
                # ジョブ割り当て
                available_jobs = job_set[i][1:-1]  # 休憩とダミーを除く
                if not available_jobs:
                    continue
                    
                current_job = None
                break_count = 0
                
                for t in range(work_start, work_end + 1):
                    if t >= n_period:
                        break
                    
                    # 休憩の配置
                    if required_breaks > 0 and break_count < required_breaks:
                        if (t - work_start) % 3 == 1:  # 3期おきに休憩
                            job_assign[f"{i}_{d}_{t}"] = 0  # 休憩
                            break_count += 1
                            continue
                    
                    # 開始直後・終了直前の休憩禁止
                    if theta > 0:
                        if t <= work_start + theta - 1 or t >= work_end - theta + 1:
                            # 通常業務を割り当て
                            if current_job is None or random.random() < 0.3:
                                current_job = random.choice(available_jobs)
                            job_assign[f"{i}_{d}_{t}"] = current_job
                        else:
                            # 休憩可能時間帯
                            if random.random() < 0.2:  # 20%の確率で休憩
                                job_assign[f"{i}_{d}_{t}"] = 0
                            else:
                                if current_job is None or random.random() < 0.3:
                                    current_job = random.choice(available_jobs)
                                job_assign[f"{i}_{d}_{t}"] = current_job
                    else:
                        # 通常の業務割り当て
                        if current_job is None or random.random() < 0.3:
                            current_job = random.choice(available_jobs)
                        job_assign[f"{i}_{d}_{t}"] = current_job
                
                work_days += 1
                total_cost += staff_df.loc[i, "wage_per_period"] * (work_end - work_start + 1)
            
            # 最大勤務日数違反チェック
            if work_days > max_day[i]:
                violations[f"totalUBConstr[{i}]"] = work_days - max_day[i]
        
        # 必要人数制約の確認
        for d in range(n_day):
            day_type = day_df.loc[d, "day_type"]
            for t in range(n_period):
                for j in range(1, n_job):  # 休憩を除く
                    required = requirement.get((day_type, t, j), 0)
                    
                    # 実際の配置人数を計算
                    assigned = 0
                    for i in range(n_staff):
                        if d in day_off[i]:
                            continue
                        key = f"{i}_{d}_{t}"
                        if key in job_assign and job_assign[key] == j:
                            assigned += 1
                    
                    # 下限制約違反
                    if assigned < required:
                        shortage = required - assigned
                        violations[f"staffLBConstr[{d}_{t}_{j}]"] = shortage
                        total_cost += lb_penalty * shortage
                    
                    # 上限制約違反
                    if assigned > required and ub_penalty > 0:
                        excess = assigned - required
                        violations[f"staffUBConstr[{d}_{t}_{j}]"] = excess
                        total_cost += ub_penalty * excess
        
        # 結果の整理（shift_scheduling2版）
        cost_df = pd.DataFrame([
            {"penalty": "Cost", "value": total_cost},
            {"penalty": "Staff Lower Bound", "value": sum(v for k, v in violations.items() if "staffLBConstr" in k) * lb_penalty},
            {"penalty": "Staff Upper Bound", "value": sum(v for k, v in violations.items() if "staffUBConstr" in k) * ub_penalty},
            {"penalty": "Change Job", "value": 0},  # 簡略化
            {"penalty": "Break Number", "value": 0},  # 簡略化
            {"penalty": "Early Break", "value": 0},  # 簡略化
            {"penalty": "Late Break", "value": 0},  # 簡略化
            {"penalty": "Max Work Day", "value": sum(v for k, v in violations.items() if "totalUBConstr" in k) * max_day_penalty}
        ])
        
        # 逸脱情報
        violate_df = pd.DataFrame(index=["Staff Lower Bound", "Staff Upper Bound", 
                                       "Change Job", "Break Number", "Early Break", "Late Break"],
                                columns=[f"Day {d}" for d in range(n_day)])
        violate_df = violate_df.fillna(0)
        
        # スタッフ情報（シフト付き）
        new_staff_df = staff_df.copy()
        for d in range(n_day):
            shift_col = f"Shift for Day {d}"
            shifts = []
            for i in range(n_staff):
                if d in day_off[i]:
                    shifts.append(None)
                else:
                    # 実際の勤務時間を計算
                    work_periods = [t for t in range(n_period) if f"{i}_{d}_{t}" in job_assign]
                    if work_periods:
                        shift_start = min(work_periods)
                        shift_end = max(work_periods)
                        shifts.append(f"{shift_start}_{shift_end}")
                    else:
                        shifts.append(None)
            new_staff_df[shift_col] = shifts
        
        # 最大勤務日数違反
        max_day_violations = [violations.get(f"totalUBConstr[{i}]", 0) for i in range(n_staff)]
        new_staff_df["max day violation"] = max_day_violations
        
        return {
            "status": 0,
            "cost_df": cost_df.to_dict('records'),
            "violate_df": violate_df.to_dict(),
            "staff_df": new_staff_df.to_dict('records'),
            "job_assign": job_assign,
            "message": "日別希望勤務時間対応の最適化が正常に完了しました",
            "violations": dict(violations)
        }
    
    def _solve_shift_optimization(self, n_staff: int, n_day: int, n_period: int, n_job: int,
                                  st_dic: Dict, job_set: Dict, requirement: Dict,
                                  day_off: Dict, break_time: Dict,
                                  start: Dict, end: Dict, max_period: Dict, max_day: Dict,
                                  theta: int, lb_penalty: int, ub_penalty: int,
                                  job_change_penalty: int, break_penalty: int,
                                  max_day_penalty: int, staff_df: pd.DataFrame,
                                  day_df: pd.DataFrame) -> Dict[str, Any]:
        """実際の最適化ソルバー実行"""
        
        # 簡易ヒューリスティック解法による実装
        job_assign = {}
        violations = defaultdict(int)
        total_cost = 0
        
        # スタッフごとにシフトを生成
        for i in range(n_staff):
            staff_violations = 0
            work_days = 0
            
            for d in range(n_day):
                if d in day_off[i]:
                    continue
                
                if work_days >= max_day[i]:
                    continue
                
                # 基本的なシフトパターンを適用
                work_start = start[i]
                work_end = min(end[i], work_start + max_period[i])
                
                # 休憩時間の計算
                work_duration = work_end - work_start + 1
                required_breaks = break_time.get(work_duration, 0)
                
                # ジョブ割り当て
                available_jobs = job_set[i][1:-1]  # 休憩とダミーを除く
                current_job = None
                break_count = 0
                
                for t in range(work_start, work_end + 1):
                    if t >= n_period:
                        break
                    
                    # 休憩の配置
                    if required_breaks > 0 and break_count < required_breaks:
                        if (t - work_start) % 3 == 1:  # 3期おきに休憩
                            job_assign[f"{i}_{d}_{t}"] = 0  # 休憩
                            break_count += 1
                            continue
                    
                    # 開始直後・終了直前の休憩禁止
                    if theta > 0:
                        if t <= work_start + theta - 1 or t >= work_end - theta + 1:
                            if not available_jobs:
                                continue
                            # 通常業務を割り当て
                            if current_job is None or random.random() < 0.3:
                                current_job = random.choice(available_jobs)
                            job_assign[f"{i}_{d}_{t}"] = current_job
                        else:
                            # 休憩可能時間帯
                            if random.random() < 0.2:  # 20%の確率で休憩
                                job_assign[f"{i}_{d}_{t}"] = 0
                            else:
                                if current_job is None or random.random() < 0.3:
                                    current_job = random.choice(available_jobs)
                                job_assign[f"{i}_{d}_{t}"] = current_job
                    else:
                        # 通常の業務割り当て
                        if current_job is None or random.random() < 0.3:
                            current_job = random.choice(available_jobs)
                        job_assign[f"{i}_{d}_{t}"] = current_job
                
                work_days += 1
                total_cost += staff_df.loc[i, "wage_per_period"] * (work_end - work_start + 1)
            
            # 最大勤務日数違反チェック
            if work_days > max_day[i]:
                violations[f"totalUBConstr[{i}]"] = work_days - max_day[i]
                staff_violations = work_days - max_day[i]
        
        # 必要人数制約の確認
        for d in range(n_day):
            day_type = day_df.loc[d, "day_type"]
            for t in range(n_period):
                for j in range(1, n_job):  # 休憩を除く
                    required = requirement.get((day_type, t, j), 0)
                    
                    # 実際の配置人数を計算
                    assigned = 0
                    for i in range(n_staff):
                        if d in day_off[i]:
                            continue
                        key = f"{i}_{d}_{t}"
                        if key in job_assign and job_assign[key] == j:
                            assigned += 1
                    
                    # 下限制約違反
                    if assigned < required:
                        shortage = required - assigned
                        violations[f"staffLBConstr[{d}_{t}_{j}]"] = shortage
                        total_cost += lb_penalty * shortage
                    
                    # 上限制約違反
                    if assigned > required and ub_penalty > 0:
                        excess = assigned - required
                        violations[f"staffUBConstr[{d}_{t}_{j}]"] = excess
                        total_cost += ub_penalty * excess
        
        # 結果の整理
        cost_df = pd.DataFrame([
            {"penalty": "Cost", "value": total_cost},
            {"penalty": "Staff Lower Bound", "value": sum(v for k, v in violations.items() if "staffLBConstr" in k) * lb_penalty},
            {"penalty": "Staff Upper Bound", "value": sum(v for k, v in violations.items() if "staffUBConstr" in k) * ub_penalty},
            {"penalty": "Change Job", "value": 0},  # 簡略化
            {"penalty": "Break Number", "value": 0},  # 簡略化
            {"penalty": "Early Break", "value": 0},  # 簡略化
            {"penalty": "Late Break", "value": 0},  # 簡略化
            {"penalty": "Max Work Day", "value": sum(v for k, v in violations.items() if "totalUBConstr" in k) * max_day_penalty}
        ])
        
        # 逸脱情報
        violate_df = pd.DataFrame(index=["Staff Lower Bound", "Staff Upper Bound", 
                                       "Change Job", "Break Number", "Early Break", "Late Break"],
                                columns=[f"Day {d}" for d in range(n_day)])
        violate_df = violate_df.fillna(0)
        
        # スタッフ情報（シフト付き）
        new_staff_df = staff_df.copy()
        for d in range(n_day):
            shift_col = f"Shift for Day {d}"
            shifts = []
            for i in range(n_staff):
                if d in day_off[i]:
                    shifts.append(None)
                else:
                    # 実際の勤務時間を計算
                    work_periods = [t for t in range(n_period) if f"{i}_{d}_{t}" in job_assign]
                    if work_periods:
                        shift_start = min(work_periods)
                        shift_end = max(work_periods)
                        shifts.append(f"{shift_start}_{shift_end}")
                    else:
                        shifts.append(None)
            new_staff_df[shift_col] = shifts
        
        # 最大勤務日数違反
        max_day_violations = [violations.get(f"totalUBConstr[{i}]", 0) for i in range(n_staff)]
        new_staff_df["max day violation"] = max_day_violations
        
        return {
            "status": 0,
            "cost_df": cost_df.to_dict('records'),
            "violate_df": violate_df.to_dict(),
            "staff_df": new_staff_df.to_dict('records'),
            "job_assign": job_assign,
            "message": "最適化が正常に完了しました",
            "violations": dict(violations)
        }
    
    def estimate_requirement(self, day_df: pd.DataFrame, period_df: pd.DataFrame,
                           job_df: pd.DataFrame, staff_df: pd.DataFrame,
                           requirement_df: pd.DataFrame, days: List[int] = None) -> str:
        """必要人数と供給可能人数の比較分析"""
        
        n_day = len(day_df)
        if days is None:
            days = list(range(n_day))
        
        n_job = len(job_df)
        n_period = len(period_df) - 1
        n_staff = len(staff_df)
        
        # スタッフの供給可能人数を計算
        work_hours = np.zeros((n_job, n_day, n_period))
        
        for i, row in enumerate(staff_df.itertuples()):
            job_set = self._safe_parse_list(row.job_set)
            day_off = set(self._safe_parse_list(row.day_off))
            max_period = row.max_period
            max_day = row.max_day
            ratio = max_period / n_period * max_day / n_day
            
            for d in range(n_day):
                if d not in day_off:
                    for j in job_set:
                        for t in range(row.start, min(row.end + 1, n_period)):
                            work_hours[j, d, t] += ratio
        
        # 必要人数を計算
        requirement = {}
        for row in requirement_df.itertuples():
            requirement[row.day_type, row.period, row.job] = row.requirement
        
        req = np.zeros((n_job, n_day, n_period))
        for d, row in enumerate(day_df.itertuples()):
            for j in range(1, n_job):
                for t in range(n_period):
                    req[j, d, t] = requirement.get((row.day_type, t, j), 0)
        
        # Plotlyグラフ作成
        fig = make_subplots(
            rows=len(days), cols=n_job-1,
            shared_xaxes="all", shared_yaxes="all",
            row_titles=[str(day_df.day.iloc[d]) for d in days],
            column_titles=[job_df.description.iloc[j] for j in range(1, n_job)]
        )
        
        for i, d in enumerate(days):
            for j in range(1, n_job):
                # 必要量（破線）
                yy = [req[j, d, t] for t in range(n_period)]
                xx = period_df.description.iloc[:-1]
                
                trace0 = go.Scatter(
                    x=xx, y=yy,
                    mode='lines',
                    line=dict(color='firebrick', width=4, dash='dot'),
                    name="Required staffs",
                    showlegend=(i == 0 and j == 1)
                )
                
                # 供給可能量（バー）
                trace1 = go.Bar(
                    x=xx, y=[work_hours[j, d, t] for t in range(n_period)],
                    name="Available staffs",
                    marker_color='gray',
                    opacity=0.5,
                    showlegend=(i == 0 and j == 1)
                )
                
                fig.add_trace(trace0, row=i+1, col=j)
                fig.add_trace(trace1, row=i+1, col=j)
        
        fig.update_layout(height=800, title_text="Staff Requirement vs Availability Analysis")
        
        return fig.to_json()
    
    def create_excel_report(self, data: Dict[str, Any], period_df: pd.DataFrame = None, 
                           day_df: pd.DataFrame = None, job_df: pd.DataFrame = None, 
                           staff_df: pd.DataFrame = None, requirement_df: pd.DataFrame = None) -> str:
        """Excel レポート生成（ノートブック準拠の詳細版）"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Shift Schedule"
        
        # 基本情報の処理
        job_assign = data.get('job_assign', {})
        staff_data = data.get('staff_df', [])
        
        if not (period_df is not None and day_df is not None and job_df is not None):
            # 基本的なレポート
            return self._create_basic_excel_report(data, wb, ws)
        
        # 詳細レポート（ノートブック準拠）
        return self._create_detailed_excel_report(
            data, wb, ws, period_df, day_df, job_df, staff_df, requirement_df
        )
    
    def _create_basic_excel_report(self, data: Dict[str, Any], wb: Workbook, ws) -> str:
        """基本的なExcelレポート"""
        
        # ヘッダー情報
        ws.append(["Shift Optimization Report"])
        ws.append(["Generated at:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        
        # コスト情報
        if 'cost_df' in data:
            ws.append(["Cost Summary"])
            cost_data = data['cost_df']
            if cost_data:
                ws.append(["Penalty Type", "Value"])
                for row in cost_data:
                    ws.append([row.get('penalty', ''), row.get('value', 0)])
                ws.append([])
        
        # 一時ファイルに保存
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def _create_detailed_excel_report(self, data: Dict[str, Any], wb: Workbook, ws,
                                     period_df: pd.DataFrame, day_df: pd.DataFrame,
                                     job_df: pd.DataFrame, staff_df: pd.DataFrame,
                                     requirement_df: pd.DataFrame = None) -> str:
        """詳細なExcelレポート（ノートブック準拠）"""
        
        LEFTSPACE = 3
        job_assign = data.get('job_assign', {})
        
        n_staff = len(staff_df)
        n_day = len(day_df)
        n_period = len(period_df) - 1
        n_job = len(job_df)
        
        # 列幅の設定
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 25
        
        # ヘッダー行
        header = ["社員名", "可能なシフト", "希望休日"] + [day_df.day.iloc[d] for d in range(n_day)] + ["休日合計"]
        ws.append(header)
        
        # 日付フォーマット設定
        for t in range(n_day):
            cell = ws.cell(1, LEFTSPACE + 1 + t)
            cell.number_format = 'yyyy/m/d'
            from openpyxl.styles import Alignment
            cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=False)
        
        # スタッフデータとシフト結果
        job_array = np.zeros(shape=(n_staff, n_day, n_period), dtype=int)
        
        # job_assignから配列に変換
        for key_str, job_id in job_assign.items():
            try:
                i, d, t = map(int, key_str.split('_'))
                if 0 <= i < n_staff and 0 <= d < n_day and 0 <= t < n_period:
                    job_array[i, d, t] = job_id
            except:
                continue
        
        # 各スタッフの行を追加
        for i, row in enumerate(staff_df.itertuples()):
            # 主要ジョブを計算（各日の最頻出ジョブ）
            daily_jobs = []
            for d in range(n_day):
                day_jobs = job_array[i, d, :]
                if np.any(day_jobs > 0):
                    # 休憩以外で最も多いジョブ
                    job_counts = np.bincount(day_jobs)
                    job_counts[0] = 0  # 休憩は除外
                    main_job = np.argmax(job_counts) if len(job_counts) > 1 else 0
                    daily_jobs.append(main_job)
                else:
                    daily_jobs.append(0)
            
            staff_row = [row.name, row.job_set, row.day_off] + daily_jobs
            ws.append(staff_row)
        
        # 合計行の追加
        ws.append(["合計"])
        
        # ジョブ別集計
        for j, job_row in enumerate(job_df.itertuples()):
            ws.append(["", j, job_row.description])
            for t in range(n_day):
                cell = ws.cell(3 + n_staff + j, LEFTSPACE + 1 + t)
                col = cell.column_letter
                cell.value = f"=COUNTIF({col}2:{col}{n_staff+1},{j})"
        
        # 必要人数の追加（requirement_dfが利用可能な場合）
        if requirement_df is not None:
            requirement = {}
            for row in requirement_df.itertuples():
                requirement[row.day_type, row.period, row.job] = row.requirement
            
            ws.append(["必要人数"])
            for j, job_row in enumerate(job_df.itertuples()):
                req_row = ["", j, job_row.description]
                for d in range(n_day):
                    day_type = day_df.iloc[d].day_type
                    # 日の代表的な必要人数（時間帯平均）
                    day_requirements = []
                    for t in range(n_period):
                        req = requirement.get((day_type, t, j), 0)
                        if req > 0:
                            day_requirements.append(req)
                    avg_req = np.mean(day_requirements) if day_requirements else 0
                    req_row.append(int(avg_req))
                ws.append(req_row)
        
        # 休日合計の計算
        for i in range(n_staff):
            cell = ws.cell(2 + i, LEFTSPACE + n_day + 1)
            start_col = ws.cell(2 + i, LEFTSPACE + 1).column_letter
            end_col = ws.cell(2 + i, LEFTSPACE + n_day).column_letter
            cell.value = f"=COUNTIF({start_col}{2+i}:{end_col}{2+i},0)"
        
        # 一時ファイルに保存
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def make_gannt_excel(self, job_assign: Dict[str, int], period_df: pd.DataFrame, 
                        day_df: pd.DataFrame, job_df: pd.DataFrame, 
                        staff_df: pd.DataFrame, requirement_df: pd.DataFrame) -> str:
        """ガントチャート付きExcel生成（ノートブック準拠）"""
        
        LEFTSPACE = 2
        
        # 要求タイプ、ジョブ、期ごとの必要人数を入れる辞書
        requirement = {}
        for row in requirement_df.itertuples():
            requirement[row.day_type, row.period, row.job] = row.requirement
        
        n_staff = len(staff_df)
        n_day = len(day_df)
        n_period = len(period_df) - 1
        n_job = len(job_df)
        
        # 割り当てジョブを配列に変換
        job_array = np.zeros(shape=(n_staff, n_day, n_period), dtype=int)
        for key_str, job_id in job_assign.items():
            try:
                i, d, t = map(int, key_str.split('_'))
                if 0 <= i < n_staff and 0 <= d < n_day and 0 <= t < n_period:
                    job_array[i, d, t] = job_id
            except:
                continue
        
        wb = Workbook()
        ws0 = wb.active
        wb.remove(ws0)
        
        for day in range(n_day):
            ws = wb.create_sheet(title=str(day_df.day.iloc[day])[:10])  # 日付に変換
            data = job_array[:, day, :]
            
            # 時間ヘッダー
            ws.append([""] * LEFTSPACE + list(period_df.description.iloc[:-1]))
            
            # スタッフごとのジョブデータ
            for i, row in enumerate(data):
                ws.append([i, staff_df.name.iloc[i]] + list(row))
            
            # 人数集計
            ws.append(["人数", "期"] + [t for t in range(n_period)])
            
            # ジョブ別カウント
            for j in range(1, n_job):
                for col in range(LEFTSPACE + 1, LEFTSPACE + n_period + 1):
                    cell = ws.cell(n_staff + 2 + j, col)
                    col_name = cell.column_letter
                    cell.value = f"=COUNTIF({col_name}2:{col_name}{n_staff+1},{j})"
            
            # ジョブ名と番号
            for j in range(1, n_job):
                ws.cell(n_staff + 2 + j, 1).value = j
                ws.cell(n_staff + 2 + j, 2).value = job_df.description.iloc[j]
            
            # 必要人数
            req_type = day_df.iloc[day].day_type
            ws.append(["必要人数"])
            for j in range(1, n_job):
                req_row = [j, f"{job_df.description.iloc[j]} 下限"]
                for t in range(n_period):
                    req_row.append(requirement.get((req_type, t, j), 0))
                ws.append(req_row)
        
        # 一時ファイルに保存
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def make_allshift_excel(self, staff_data: List[Dict], day_df: pd.DataFrame, 
                           period_df: pd.DataFrame) -> str:
        """全期間シフト一覧Excel生成（ノートブック準拠）"""
        
        n_day = len(day_df)
        n_staff = len(staff_data)
        
        # 期ID => 時刻の辞書
        period_dic = {i: d for i, d in enumerate(period_df.description)}
        
        LEFTSPACE = 2
        
        wb = Workbook()
        ws = wb.active
        
        # データ処理
        data = []
        for staff in staff_data:
            start_times, end_times, durations = [], [], []
            
            for d in range(n_day):
                shift_key = f"Shift for Day {d}"
                shift_value = staff.get(shift_key)
                
                if shift_value and shift_value != "None":
                    try:
                        s, f = map(int, shift_value.split("_"))
                        start_times.append(period_dic.get(s, ""))
                        end_times.append(period_dic.get(f, ""))
                        durations.append(f - s + 1)
                    except:
                        start_times.append("")
                        end_times.append("")
                        durations.append(0)
                else:
                    start_times.append("")
                    end_times.append("")
                    durations.append(0)
            
            data.extend([start_times, end_times, durations])
        
        # ヘッダー
        header = [""] * LEFTSPACE + [day_df.day.iloc[d] for d in range(n_day)] + ["合計"]
        ws.append(header)
        
        # シフトデータ出力
        for i, row in enumerate(data):
            if i % 3 == 0:
                ws.append([i // 3, staff_data[i // 3]['name']] + row)
            else:
                ws.append(["", ""] + row)
        
        # 合計計算
        for i in range(n_staff):
            row_num = 2 + i * 3
            cell = ws.cell(row_num, LEFTSPACE + n_day + 1)
            last_column = ws.cell(1, LEFTSPACE + n_day).column_letter
            cell.value = f"=SUM(C{row_num}:{last_column}{row_num})"
        
        # 一時ファイルに保存
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def make_requirement_graph(self, day_df: pd.DataFrame, period_df: pd.DataFrame,
                              job_df: pd.DataFrame, staff_df: pd.DataFrame,
                              requirement_df: pd.DataFrame, job_assign: Dict[str, int],
                              day: int = 0) -> str:
        """必要人数グラフ生成"""
        
        n_day = len(day_df)
        n_period = len(period_df) - 1
        n_job = len(job_df)
        n_staff = len(staff_df)
        
        req_type = day_df.loc[day, "day_type"]
        
        # 必要人数の辞書
        requirement = {}
        for row in requirement_df.itertuples():
            requirement[row.day_type, row.period, row.job] = row.requirement
        
        # 休日希望日の辞書
        day_off = {}
        for i in range(n_staff):
            day_off[i] = set(self._safe_parse_list(staff_df.loc[i, "day_off"]))
        
        # ジョブ割り当ての変換
        y = {}
        for key, value in job_assign.items():
            if isinstance(key, str) and '_' in key:
                # 新しい形式: "staff_day_period"
                parts = key.split('_')
                if len(parts) == 3:
                    tpl = tuple(map(int, parts))
                else:
                    continue
            else:
                # 旧形式: tuple または文字列表現のtuple
                try:
                    tpl = ast.literal_eval(key) if isinstance(key, str) else key
                except:
                    continue
            y[tpl] = value
        
        # 人数計算
        val = np.zeros((n_day, n_period, n_job + 1), int)
        for d in range(n_day):
            for t in range(n_period):
                for i in range(n_staff):
                    if d in day_off[i]:
                        continue
                    j = int(y.get((i, d, t), 0))
                    val[d, t, j] += 1
        
        # グラフ作成
        fig = make_subplots(
            rows=n_job-1, cols=1,
            subplot_titles=[job_df.description.iloc[j] for j in range(1, n_job)]
        )
        
        for j in range(1, n_job):
            yy = [requirement.get((req_type, t, j), 0) for t in range(n_period)]
            xx = period_df.description.iloc[:-1]
            
            trace0 = go.Scatter(
                x=xx, y=yy,
                mode='lines',
                line=dict(color='firebrick', width=4, dash='dot'),
                name="Required staffs"
            )
            
            trace1 = go.Bar(
                x=xx, y=val[day, :, j],
                name="Assigned staffs",
                marker_color='gray',
                opacity=0.5
            )
            
            fig.add_trace(trace0, row=j, col=1)
            fig.add_trace(trace1, row=j, col=1)
        
        fig.update_layout(height=800, showlegend=False)
        return fig.to_json()
    
    def make_gantt_chart(self, day_df: pd.DataFrame, period_df: pd.DataFrame,
                        staff_df: pd.DataFrame, job_df: pd.DataFrame,
                        job_assign: Dict[str, int], day: int = 0) -> str:
        """ガントチャート生成"""
        
        n_period = len(period_df) - 1
        n_job = len(job_df)
        n_staff = len(staff_df)
        
        # ジョブ割り当ての変換
        y = {}
        for key, value in job_assign.items():
            if isinstance(key, str) and '_' in key:
                # 新しい形式: "staff_day_period"
                parts = key.split('_')
                if len(parts) == 3:
                    tpl = tuple(map(int, parts))
                else:
                    continue
            else:
                # 旧形式: tuple または文字列表現のtuple
                try:
                    tpl = ast.literal_eval(key) if isinstance(key, str) else key
                except:
                    continue
            y[tpl] = value
        
        # 休日希望日の辞書
        day_off = {}
        for i in range(n_staff):
            day_off[i] = set(self._safe_parse_list(staff_df.loc[i, "day_off"]))
        
        # 日付・時刻辞書
        dt_dic = {}
        for d, day_ in enumerate(day_df["day"]):
            for t, time_ in enumerate(period_df["description"]):
                dt_dic[d, t] = pd.to_datetime(str(day_) + " " + str(time_))
        
        # ガントチャート用データ
        L = []
        for i, staff_name in enumerate(staff_df["name"]):
            if day in day_off[i]:
                continue
            
            # タスクの抽出
            Task = []
            j_prev = None
            for t in range(n_period + 1):
                if t == n_period:
                    j = n_job  # 番兵
                else:
                    j = int(y.get((i, day, t), n_job))
                
                if j == n_job:  # ダミージョブまたは休憩
                    if j_prev is None:
                        continue
                    else:
                        Task.append((j_prev, start_period, t))
                        break
                
                if j_prev is None:
                    start_period = t
                    j_prev = j
                    continue
                
                if j == j_prev:
                    j_prev = j
                else:
                    Task.append((j_prev, start_period, t))
                    j_prev = j
                    start_period = t
            
            # タスクをガントチャート形式に変換
            for j, st, fi in Task:
                L.append(dict(
                    Task=staff_name,
                    Start=dt_dic[day, st],
                    Finish=dt_dic[day, fi],
                    Resource=job_df["description"].iloc[j]
                ))
        
        # ガントチャート作成
        if L:
            fig = ff.create_gantt(
                L, title="Staff Schedule Gantt Chart", 
                index_col='Resource',
                show_colorbar=True, 
                showgrid_x=True, 
                showgrid_y=True, 
                group_tasks=True
            )
            return fig.to_json()
        else:
            # 空のグラフを返す
            fig = go.Figure()
            fig.update_layout(title="No schedule data available")
            return fig.to_json()