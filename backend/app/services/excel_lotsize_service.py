import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple, Any, Optional, Union
from collections import defaultdict
import os
import warnings
warnings.filterwarnings('ignore')


class ExcelLotsizeService:
    """
    Excel Integration Service for Lot Size Optimization
    Exact implementation from 11lotsize.ipynb notebook for Excel I/O operations
    
    Supports:
    - Master data generation for Excel templates
    - Reading optimization data from Excel files
    - Outputting optimization results to Excel
    - Multi-mode optimization Excel handling
    """
    
    def __init__(self):
        self.default_sheet_names = {
            'items': 'Item Master',
            'processes': 'Process Master', 
            'bom': 'BOM',
            'resources': 'Resource',
            'orders': 'Order Master',
            'resource_detail': 'Resource Detail'
        }
    
    def generate_lotsize_master(self, prod_df: pd.DataFrame, 
                              production_df: pd.DataFrame,
                              bom_df: pd.DataFrame,
                              demand: np.ndarray,
                              resource_df: pd.DataFrame,
                              filename: str = "lotsize_master.xlsx") -> str:
        """
        ロットサイズ最適化用のマスターExcelファイルを生成
        Exact implementation from notebook
        
        Args:
            prod_df: 品目データフレーム
            production_df: 生産情報データフレーム
            bom_df: 部品展開表データフレーム
            demand: 需要配列
            resource_df: 資源データフレーム
            filename: 出力ファイル名
            
        Returns:
            生成されたファイルのパス
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate Item Master sheet
        self._add_item_master_sheet(wb, prod_df)
        
        # Generate Process Master sheet
        self._add_process_master_sheet(wb, production_df)
        
        # Generate BOM sheet
        if bom_df is not None and len(bom_df) > 0:
            self._add_bom_sheet(wb, bom_df)
        
        # Generate Resource sheet
        if resource_df is not None and len(resource_df) > 0:
            self._add_resource_sheet(wb, resource_df)
        
        # Generate Order Master sheet from demand
        self._add_order_master_sheet(wb, demand, prod_df)
        
        # Save workbook
        wb.save(filename)
        return filename
    
    def generate_item_master(self, prod_df: pd.DataFrame) -> pd.DataFrame:
        """
        品目マスタを生成
        Exact implementation from notebook
        """
        item_master = prod_df.copy()
        
        # Standard columns for item master
        required_cols = ['name', 'holding_cost', 'setup_cost', 'target_inventory']
        
        # Add missing columns with default values
        for col in required_cols:
            if col not in item_master.columns:
                if col == 'holding_cost':
                    item_master[col] = 1.0
                elif col == 'setup_cost':
                    item_master[col] = 100.0
                elif col == 'target_inventory':
                    item_master[col] = 1000.0
                else:
                    item_master[col] = ''
        
        return item_master[required_cols]
    
    def generate_process_master(self, production_df: pd.DataFrame) -> pd.DataFrame:
        """
        プロセスマスタを生成
        Exact implementation from notebook
        """
        process_master = production_df.copy()
        
        # Standard columns for process master
        required_cols = ['name', 'process', 'processing_time', 'setup_time']
        
        # Add missing columns with default values
        for col in required_cols:
            if col not in process_master.columns:
                if col == 'processing_time':
                    process_master[col] = 1.0
                elif col == 'setup_time':
                    process_master[col] = 60.0
                else:
                    process_master[col] = ''
        
        return process_master[required_cols]
    
    def add_bom_resource_sheets(self, wb: Workbook, bom_df: pd.DataFrame, 
                              resource_df: pd.DataFrame) -> None:
        """
        BOMと資源シートをワークブックに追加
        Exact implementation from notebook
        """
        if bom_df is not None and len(bom_df) > 0:
            self._add_bom_sheet(wb, bom_df)
        
        if resource_df is not None and len(resource_df) > 0:
            self._add_resource_sheet(wb, resource_df)
    
    def generate_order_master(self, demand: np.ndarray, prod_df: pd.DataFrame) -> pd.DataFrame:
        """
        注文マスタを需要配列から生成
        Exact implementation from notebook
        """
        n_items, n_periods = demand.shape
        
        order_data = []
        for item_idx in range(n_items):
            item_name = prod_df.iloc[item_idx]['name'] if 'name' in prod_df.columns else f"Item_{item_idx}"
            
            for period in range(n_periods):
                if demand[item_idx, period] > 0:
                    order_data.append({
                        'item': item_name,
                        'period': period + 1,
                        'demand': demand[item_idx, period],
                        'due_date': period + 1
                    })
        
        return pd.DataFrame(order_data)
    
    def add_detailed_resource_sheet(self, wb: Workbook, resource_df: pd.DataFrame,
                                  detailed_resource_data: Dict[str, Any]) -> None:
        """
        詳細資源シートを追加
        Exact implementation from notebook
        """
        ws = wb.create_sheet(self.default_sheet_names['resource_detail'])
        
        # Add headers
        headers = ['Resource', 'Period', 'Capacity', 'Usage', 'Utilization']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Add data
        row_idx = 2
        for resource_name, data in detailed_resource_data.items():
            for period, info in enumerate(data):
                ws.cell(row=row_idx, column=1, value=resource_name)
                ws.cell(row=row_idx, column=2, value=period + 1)
                ws.cell(row=row_idx, column=3, value=info.get('capacity', 0))
                ws.cell(row=row_idx, column=4, value=info.get('usage', 0))
                ws.cell(row=row_idx, column=5, value=info.get('utilization', 0))
                row_idx += 1
    
    def read_dfs_from_excel_lot(self, filename: str) -> Tuple[pd.DataFrame, pd.DataFrame, 
                                                            pd.DataFrame, pd.DataFrame, 
                                                            pd.DataFrame]:
        """
        ロットサイズ最適化用Excelファイルからデータフレームを読み込み
        Exact implementation from notebook
        
        Returns:
            prod_df, production_df, bom_df, resource_df, order_df
        """
        try:
            # Read each sheet
            prod_df = pd.read_excel(filename, sheet_name=self.default_sheet_names['items'])
            production_df = pd.read_excel(filename, sheet_name=self.default_sheet_names['processes'])
            
            # Optional sheets
            try:
                bom_df = pd.read_excel(filename, sheet_name=self.default_sheet_names['bom'])
            except:
                bom_df = pd.DataFrame()
            
            try:
                resource_df = pd.read_excel(filename, sheet_name=self.default_sheet_names['resources'])
            except:
                resource_df = pd.DataFrame()
            
            try:
                order_df = pd.read_excel(filename, sheet_name=self.default_sheet_names['orders'])
            except:
                order_df = pd.DataFrame()
            
            return prod_df, production_df, bom_df, resource_df, order_df
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file {filename}: {str(e)}")
    
    def get_resource_ub(self, resource_df: pd.DataFrame, period: int) -> Dict[str, float]:
        """
        指定期間の資源上限を取得
        Exact implementation from notebook
        """
        resource_ub = {}
        
        if resource_df is not None and len(resource_df) > 0:
            period_data = resource_df[resource_df['period'] == period]
            for _, row in period_data.iterrows():
                resource_ub[row['name']] = row['capacity']
        
        return resource_ub
    
    def extract_fix_info(self, wb: Workbook, sheet_name: str = "Fix Info") -> Dict[str, Any]:
        """
        固定情報シートから変数固定情報を抽出
        Exact implementation from notebook for rolling horizon optimization
        """
        fix_info = {'fixed_variables': {}, 'fix_periods': 0}
        
        try:
            ws = wb[sheet_name]
            
            # Read fix periods
            fix_periods_cell = ws['B1']
            if fix_periods_cell.value:
                fix_info['fix_periods'] = int(fix_periods_cell.value)
            
            # Read fixed variables (starting from row 3)
            row_idx = 3
            while ws[f'A{row_idx}'].value:
                item = ws[f'A{row_idx}'].value
                period = int(ws[f'B{row_idx}'].value)
                var_type = ws[f'C{row_idx}'].value  # 'x', 'I', 'y'
                value = ws[f'D{row_idx}'].value
                
                if item not in fix_info['fixed_variables']:
                    fix_info['fixed_variables'][item] = {}
                if period not in fix_info['fixed_variables'][item]:
                    fix_info['fixed_variables'][item][period] = {}
                
                fix_info['fixed_variables'][item][period][var_type] = value
                row_idx += 1
                
        except Exception as e:
            print(f"Warning: Could not read fix info: {str(e)}")
        
        return fix_info
    
    def lot_output_excel(self, model: Any, prod_df: pd.DataFrame, 
                        production_df: pd.DataFrame, T: int,
                        filename: str = "lotsize_result.xlsx") -> str:
        """
        ロットサイズ最適化結果をExcelに出力
        Exact implementation from notebook
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Production schedule sheet
        self._add_production_schedule_sheet(wb, model, prod_df, T)
        
        # Inventory level sheet
        self._add_inventory_sheet(wb, model, prod_df, T)
        
        # Setup schedule sheet
        self._add_setup_schedule_sheet(wb, model, prod_df, T)
        
        # Cost summary sheet
        self._add_cost_summary_sheet(wb, model, prod_df, production_df, T)
        
        wb.save(filename)
        return filename
    
    def show_result_for_multimode_lotsizing(self, model: Any, prod_df: pd.DataFrame,
                                          production_df: pd.DataFrame, T: int,
                                          modes: List[str]) -> Dict[str, pd.DataFrame]:
        """
        マルチモードロットサイズ最適化結果を表示用に整理
        Exact implementation from notebook
        """
        results = {}
        
        # Production by mode
        prod_by_mode = []
        for mode in modes:
            mode_data = []
            for item_idx, item_name in enumerate(prod_df['name']):
                for t in range(T):
                    try:
                        if hasattr(model, 'x') and (t, item_name, mode) in model.x:
                            value = model.x[t, item_name, mode].x if hasattr(model.x[t, item_name, mode], 'x') else model.x[t, item_name, mode]
                            if value > 0.01:
                                mode_data.append({
                                    'Item': item_name,
                                    'Period': t + 1,
                                    'Mode': mode,
                                    'Quantity': value
                                })
                    except:
                        continue
            
            if mode_data:
                results[f'Production_Mode_{mode}'] = pd.DataFrame(mode_data)
        
        # Setup schedule by mode
        setup_by_mode = []
        for mode in modes:
            mode_data = []
            for item_idx, item_name in enumerate(prod_df['name']):
                for t in range(T):
                    try:
                        if hasattr(model, 'y') and (t, item_name, mode) in model.y:
                            value = model.y[t, item_name, mode].x if hasattr(model.y[t, item_name, mode], 'x') else model.y[t, item_name, mode]
                            if value > 0.5:
                                mode_data.append({
                                    'Item': item_name,
                                    'Period': t + 1,
                                    'Mode': mode,
                                    'Setup': 1
                                })
                    except:
                        continue
            
            if mode_data:
                results[f'Setup_Mode_{mode}'] = pd.DataFrame(mode_data)
        
        # Resource utilization by mode
        if hasattr(model, 'resource_usage'):
            resource_util = []
            for mode in modes:
                for resource_name in production_df['resource'].unique():
                    for t in range(T):
                        try:
                            usage = model.resource_usage.get((resource_name, t, mode), 0)
                            if usage > 0.01:
                                resource_util.append({
                                    'Resource': resource_name,
                                    'Period': t + 1,
                                    'Mode': mode,
                                    'Usage': usage
                                })
                        except:
                            continue
            
            if resource_util:
                results['Resource_Utilization'] = pd.DataFrame(resource_util)
        
        return results
    
    def _add_item_master_sheet(self, wb: Workbook, prod_df: pd.DataFrame) -> None:
        """品目マスタシートを追加"""
        ws = wb.create_sheet(self.default_sheet_names['items'])
        
        item_master = self.generate_item_master(prod_df)
        
        # Add data to sheet
        for r in dataframe_to_rows(item_master, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    def _add_process_master_sheet(self, wb: Workbook, production_df: pd.DataFrame) -> None:
        """プロセスマスタシートを追加"""
        ws = wb.create_sheet(self.default_sheet_names['processes'])
        
        process_master = self.generate_process_master(production_df)
        
        # Add data to sheet
        for r in dataframe_to_rows(process_master, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    def _add_bom_sheet(self, wb: Workbook, bom_df: pd.DataFrame) -> None:
        """BOMシートを追加"""
        ws = wb.create_sheet(self.default_sheet_names['bom'])
        
        # Add data to sheet
        for r in dataframe_to_rows(bom_df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    def _add_resource_sheet(self, wb: Workbook, resource_df: pd.DataFrame) -> None:
        """資源シートを追加"""
        ws = wb.create_sheet(self.default_sheet_names['resources'])
        
        # Add data to sheet
        for r in dataframe_to_rows(resource_df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    def _add_order_master_sheet(self, wb: Workbook, demand: np.ndarray, 
                              prod_df: pd.DataFrame) -> None:
        """注文マスタシートを追加"""
        ws = wb.create_sheet(self.default_sheet_names['orders'])
        
        order_master = self.generate_order_master(demand, prod_df)
        
        # Add data to sheet
        for r in dataframe_to_rows(order_master, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    def _add_production_schedule_sheet(self, wb: Workbook, model: Any, 
                                     prod_df: pd.DataFrame, T: int) -> None:
        """生産スケジュールシートを追加"""
        ws = wb.create_sheet("Production Schedule")
        
        # Headers
        headers = ['Item', 'Period', 'Production Quantity']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data
        row_idx = 2
        for item_idx, item_name in enumerate(prod_df['name']):
            for t in range(T):
                try:
                    if hasattr(model, 'x') and (t, item_name) in model.x:
                        value = model.x[t, item_name].x if hasattr(model.x[t, item_name], 'x') else model.x[t, item_name]
                        if value > 0.01:
                            ws.cell(row=row_idx, column=1, value=item_name)
                            ws.cell(row=row_idx, column=2, value=t + 1)
                            ws.cell(row=row_idx, column=3, value=round(value, 2))
                            row_idx += 1
                except:
                    continue
    
    def _add_inventory_sheet(self, wb: Workbook, model: Any, 
                           prod_df: pd.DataFrame, T: int) -> None:
        """在庫レベルシートを追加"""
        ws = wb.create_sheet("Inventory Levels")
        
        # Headers
        headers = ['Item', 'Period', 'Inventory Level']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data
        row_idx = 2
        for item_idx, item_name in enumerate(prod_df['name']):
            for t in range(T):
                try:
                    if hasattr(model, 'I') and (t, item_name) in model.I:
                        value = model.I[t, item_name].x if hasattr(model.I[t, item_name], 'x') else model.I[t, item_name]
                        ws.cell(row=row_idx, column=1, value=item_name)
                        ws.cell(row=row_idx, column=2, value=t + 1)
                        ws.cell(row=row_idx, column=3, value=round(value, 2))
                        row_idx += 1
                except:
                    continue
    
    def _add_setup_schedule_sheet(self, wb: Workbook, model: Any, 
                                prod_df: pd.DataFrame, T: int) -> None:
        """段取りスケジュールシートを追加"""
        ws = wb.create_sheet("Setup Schedule")
        
        # Headers
        headers = ['Item', 'Period', 'Setup']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Data
        row_idx = 2
        for item_idx, item_name in enumerate(prod_df['name']):
            for t in range(T):
                try:
                    if hasattr(model, 'y') and (t, item_name) in model.y:
                        value = model.y[t, item_name].x if hasattr(model.y[t, item_name], 'x') else model.y[t, item_name]
                        if value > 0.5:
                            ws.cell(row=row_idx, column=1, value=item_name)
                            ws.cell(row=row_idx, column=2, value=t + 1)
                            ws.cell(row=row_idx, column=3, value=1)
                            row_idx += 1
                except:
                    continue
    
    def _add_cost_summary_sheet(self, wb: Workbook, model: Any, 
                              prod_df: pd.DataFrame, production_df: pd.DataFrame, 
                              T: int) -> None:
        """コスト要約シートを追加"""
        ws = wb.create_sheet("Cost Summary")
        
        # Headers
        ws.cell(row=1, column=1, value="Cost Component")
        ws.cell(row=1, column=2, value="Total Cost")
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        
        # Calculate costs
        total_holding_cost = 0
        total_setup_cost = 0
        
        for item_idx, item_name in enumerate(prod_df['name']):
            holding_cost = prod_df.iloc[item_idx].get('holding_cost', 1.0)
            setup_cost = prod_df.iloc[item_idx].get('setup_cost', 100.0)
            
            for t in range(T):
                try:
                    # Inventory holding cost
                    if hasattr(model, 'I') and (t, item_name) in model.I:
                        inventory = model.I[t, item_name].x if hasattr(model.I[t, item_name], 'x') else model.I[t, item_name]
                        total_holding_cost += holding_cost * inventory
                    
                    # Setup cost
                    if hasattr(model, 'y') and (t, item_name) in model.y:
                        setup = model.y[t, item_name].x if hasattr(model.y[t, item_name], 'x') else model.y[t, item_name]
                        total_setup_cost += setup_cost * setup
                except:
                    continue
        
        # Add cost data
        ws.cell(row=2, column=1, value="Inventory Holding Cost")
        ws.cell(row=2, column=2, value=round(total_holding_cost, 2))
        
        ws.cell(row=3, column=1, value="Setup Cost")
        ws.cell(row=3, column=2, value=round(total_setup_cost, 2))
        
        ws.cell(row=4, column=1, value="Total Cost")
        ws.cell(row=4, column=2, value=round(total_holding_cost + total_setup_cost, 2))
        ws.cell(row=4, column=1).font = Font(bold=True)
        ws.cell(row=4, column=2).font = Font(bold=True)
    
    def get_service_info(self) -> Dict[str, Any]:
        """
        サービス情報を取得
        """
        return {
            "service_name": "Excel Lotsize Integration Service",
            "description": "Excel integration for lot size optimization system",
            "features": [
                "Master data generation (items, processes, BOM, resources, orders)",
                "Excel template creation for optimization input",
                "Result export to formatted Excel files",
                "Multi-mode optimization result handling",
                "Rolling horizon optimization support with variable fixing"
            ],
            "supported_formats": ["xlsx", "xls"],
            "default_sheets": self.default_sheet_names
        }